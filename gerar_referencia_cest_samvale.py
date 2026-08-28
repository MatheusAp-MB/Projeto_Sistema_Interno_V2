"""
Gera a planilha de referencia EAN/NCM/CEST para a Samvale, a partir do JSON
de retorno da API Sysemp (XML_Manifesto_NF_notas_mais_recentes_por_produto.json).

Gera 2 abas:
  - EAN_NCM_CEST: 1 linha por registro do JSON, plana (EAN | NCM | CEST).
  - Arvore_NCM_CEST_EAN: agrupamento hierarquico NCM -> CEST -> EAN, com
    celulas REALMENTE mescladas (nao indentacao/outline) e os numeros reais
    das NFs (campo "NR NF") por EAN, nao so uma contagem.

Rodar localmente:
    python gerar_referencia_cest_samvale.py
"""

import json
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

CAMINHO_JSON = Path(
    "integracao_sysemp/retorno_api/dados_impostos_xml_entrada/samvale/"
    "XML_Manifesto_NF_notas_mais_recentes_por_produto.json"
)
CAMINHO_SAIDA = Path("Referencia_EAN_NCM_CEST_Samvale.xlsx")

# Chaves reais confirmadas no JSON (labels em portugues, com espaco/acento --
# nao snake_case). Lista de candidatas por campo, na ordem de preferencia.
CHAVES_EAN = ["Código Barras"]
CHAVES_NCM = ["NCM Cadastro", "NCM XML"]
CHAVES_CEST = ["CEST"]
CHAVES_NF = ["NR NF"]

_AUSENTE = object()

def normalizar_valor(valor):
    """Normaliza 1 valor bruto do JSON: string tem espaços nas pontas
    removidos; se sobrar string vazia, vira None. Sem isso, um CEST vindo
    como "" (string vazia) NÃO é igual a None pro Python, e a árvore criava
    2 grupos "sem CEST" diferentes (1 rotulado, 1 em branco) pro mesmo NCM --
    bug encontrado rodando com dado real. Aplicado em EAN/NCM/CEST/NR NF pra
    não deixar a mesma armadilha acontecer em outro campo."""
    if isinstance(valor, str):
        valor = valor.strip()
        return valor or None
    return valor

def valor_da_primeira_chave_presente(registro: dict, chaves: list[str]):
    """Retorna (chave_usada, valor) da 1a chave da lista que existir no
    registro. Se nenhuma existir, retorna (None, _AUSENTE) -- nunca confunde
    "chave nao existe" com "chave existe mas o valor e None" (CEST nulo e
    realidade de negocio pra boa parte dos produtos, nao bug)."""
    for chave in chaves:
        if chave in registro:
            return chave, registro[chave]
    return None, _AUSENTE


def carregar_registros(caminho: Path) -> list[dict]:
    with caminho.open("r", encoding="utf-8") as arquivo:
        dados = json.load(arquivo)
    if isinstance(dados, list):
        return dados
    if isinstance(dados, dict):
        for chave_candidata in ("registros", "dados", "itens"):
            if isinstance(dados.get(chave_candidata), list):
                return dados[chave_candidata]
        raise ValueError(
            "JSON carregado e um dict sem uma lista de registros reconhecivel "
            "('registros'/'dados'/'itens'). Confere a estrutura real do arquivo."
        )
    raise ValueError(f"Formato inesperado no JSON: {type(dados)}")


def diagnosticar_estrutura(registros: list[dict]) -> None:
    print(f"Total de registros no JSON: {len(registros)}")
    if not registros:
        return
    primeiro = registros[0]
    print(f"Chaves do 1º registro: {sorted(primeiro.keys())}")

    for rotulo, chaves in (
        ("EAN", CHAVES_EAN),
        ("NCM", CHAVES_NCM),
        ("CEST", CHAVES_CEST),
        ("NF", CHAVES_NF),
    ):
        chave_usada, _ = valor_da_primeira_chave_presente(primeiro, chaves)
        if chave_usada is None:
            print(f"ATENÇÃO: nenhuma das chaves candidatas de {rotulo} {chaves} foi encontrada no 1º registro.")


def extrair_linhas_referencia(registros: list[dict]) -> list[tuple]:
    """Retorna 1 tupla (ean, ncm, cest, nr_nf) por registro do JSON -- sem
    dedupe, sem agrupamento; e a base plana usada pelas 2 abas."""
    linhas = []
    for registro in registros:
        _, ean = valor_da_primeira_chave_presente(registro, CHAVES_EAN)
        _, ncm = valor_da_primeira_chave_presente(registro, CHAVES_NCM)
        _, cest = valor_da_primeira_chave_presente(registro, CHAVES_CEST)
        _, nr_nf = valor_da_primeira_chave_presente(registro, CHAVES_NF)

        linhas.append((
            normalizar_valor(None if ean is _AUSENTE else ean),
            normalizar_valor(None if ncm is _AUSENTE else ncm),
            normalizar_valor(None if cest is _AUSENTE else cest),
            normalizar_valor(None if nr_nf is _AUSENTE else nr_nf),
        ))
        
    return linhas


def escrever_aba_plana(workbook: Workbook, linhas_referencia: list[tuple]) -> None:
    aba = workbook.active
    aba.title = "EAN_NCM_CEST"

    aba.append(["EAN", "NCM", "CEST"])
    for celula in aba[1]:
        celula.font = Font(bold=True)

    for ean, ncm, cest, _nr_nf in linhas_referencia:
        aba.append([ean, ncm, cest])

    for indice, largura in enumerate([18, 14, 14], start=1):
        aba.column_dimensions[get_column_letter(indice)].width = largura
    aba.freeze_panes = "A2"

    # diagnostico de agrupamento (so imprime no console, nao altera a aba)
    ncms_distintos = {ncm for _, ncm, _, _ in linhas_referencia if ncm}
    sem_cest = sum(1 for _, _, cest, _ in linhas_referencia if cest is None)

    ncm_para_cests = defaultdict(set)
    for _, ncm, cest, _ in linhas_referencia:
        if ncm:
            ncm_para_cests[ncm].add(cest)
    ncms_com_mais_de_1_cest = sum(1 for cests in ncm_para_cests.values() if len(cests) > 1)

    print(f"Linhas gravadas: {len(linhas_referencia)}")
    print(f"Registros sem CEST encontrado: {sem_cest}")
    print(f"NCMs distintos: {len(ncms_distintos)}")
    print(f"NCMs com mais de 1 CEST diferente: {ncms_com_mais_de_1_cest}")


def construir_arvore(linhas_referencia: list[tuple]) -> dict:
    """Monta NCM -> CEST -> EAN -> conjunto de NR NF, a partir das linhas
    planas. CEST=None vira o ramo "(sem CEST)" -- nunca e descartado. EAN e
    sempre deduplicado dentro do grupo NCM+CEST; os NR NF daquele EAN dentro
    daquele grupo tambem sao deduplicados (um mesmo EAN pode ter chegado em
    mais de 1 nota)."""
    arvore = defaultdict(lambda: defaultdict(lambda: defaultdict(set)))

    for ean, ncm, cest, nr_nf in linhas_referencia:
        if not ncm or not ean:
            continue  # sem NCM ou sem EAN nao da pra agrupar -> fora da arvore
        arvore[ncm][cest][ean].add(nr_nf)

    return arvore


def _ordenar_nfs(conjunto_nfs: set) -> list:
    """Ordena os NR NF de 1 EAN dentro do grupo -- numerico quando possivel,
    senao texto -- e descarta valores None (NF ausente naquele registro)."""
    valores = [nf for nf in conjunto_nfs if nf is not None]

    def chave_ordenacao(valor):
        texto = str(valor)
        return (0, int(texto)) if texto.isdigit() else (1, texto)

    return sorted(valores, key=chave_ordenacao)


COR_BORDA_GRUPO_NCM = "FF2B4A63"
COR_BORDA_GRUPO_CEST = "FF8FA6BB"
COR_CABECALHO = "FFDBE6F3"
COR_COLUNA_NCM = "FFDBE6F3"
COR_COLUNA_CEST = "FFE4F0E7"
COR_COLUNA_CEST_VAZIO = "FFF1F1EE"


def escrever_aba_arvore(workbook: Workbook, arvore: dict) -> None:
    """Escreve a aba Arvore_NCM_CEST_EAN com celulas REALMENTE mescladas
    (mockup aprovado): o NCM ocupa 1 celula mesclada cobrindo todas as linhas
    de todos os CESTs dele; o CEST ocupa 1 celula mesclada cobrindo so as
    linhas dos EANs dele, dentro do bloco do NCM; EAN e sempre 1 linha propria
    (a folha da arvore, nunca mesclada); NF(s) lista os numeros reais de nota
    fiscal daquele EAN dentro daquele grupo NCM+CEST."""
    aba = workbook.create_sheet("Arvore_NCM_CEST_EAN")

    aba.append(["NCM", "CEST", "EAN", "NF(s)"])
    preenchimento_cabecalho = PatternFill("solid", fgColor=COR_CABECALHO)
    for celula in aba[1]:
        celula.font = Font(bold=True)
        celula.fill = preenchimento_cabecalho
        celula.alignment = Alignment(horizontal="left", vertical="center")

    preenchimento_ncm = PatternFill("solid", fgColor=COR_COLUNA_NCM)
    preenchimento_cest = PatternFill("solid", fgColor=COR_COLUNA_CEST)
    preenchimento_cest_vazio = PatternFill("solid", fgColor=COR_COLUNA_CEST_VAZIO)

    borda_grupo_ncm = Border(top=Side(style="medium", color=COR_BORDA_GRUPO_NCM))
    borda_grupo_cest = Border(top=Side(style="thin", color=COR_BORDA_GRUPO_CEST))

    linha_atual = 2  # linha 1 e o cabecalho

    for ncm in sorted(arvore.keys()):
        cests_do_ncm = arvore[ncm]
        # "(sem CEST)" sempre por ultimo dentro do NCM, so pra ficar previsivel
        cests_ordenados = sorted(cests_do_ncm.keys(), key=lambda c: (c is None, c))

        linha_inicio_ncm = linha_atual
        primeira_linha_do_ncm = True

        for cest in cests_ordenados:
            eans_do_cest = cests_do_ncm[cest]
            eans_ordenados = sorted(eans_do_cest.keys())

            linha_inicio_cest = linha_atual
            primeira_linha_do_cest = True

            for ean in eans_ordenados:
                texto_nfs = ", ".join(str(nf) for nf in _ordenar_nfs(eans_do_cest[ean]))

                aba.cell(row=linha_atual, column=3, value=ean).font = Font(name="Consolas")
                celula_nf = aba.cell(row=linha_atual, column=4, value=texto_nfs)
                celula_nf.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                celula_nf.font = Font(name="Consolas", size=10)

                if primeira_linha_do_ncm:
                    for coluna in (1, 2, 3, 4):
                        aba.cell(row=linha_atual, column=coluna).border = borda_grupo_ncm
                elif primeira_linha_do_cest:
                    for coluna in (2, 3, 4):
                        aba.cell(row=linha_atual, column=coluna).border = borda_grupo_cest

                primeira_linha_do_ncm = False
                primeira_linha_do_cest = False
                linha_atual += 1

            linha_fim_cest = linha_atual - 1
            celula_cest = aba.cell(row=linha_inicio_cest, column=2)
            eh_vazio = cest is None
            celula_cest.value = "(sem CEST)" if eh_vazio else cest
            celula_cest.alignment = Alignment(horizontal="center", vertical="center")
            celula_cest.font = Font(italic=True, color="FF6B7280") if eh_vazio else Font(bold=True, name="Consolas")
            celula_cest.fill = preenchimento_cest_vazio if eh_vazio else preenchimento_cest
            if linha_fim_cest > linha_inicio_cest:
                aba.merge_cells(start_row=linha_inicio_cest, start_column=2, end_row=linha_fim_cest, end_column=2)

        linha_fim_ncm = linha_atual - 1
        celula_ncm = aba.cell(row=linha_inicio_ncm, column=1)
        celula_ncm.value = ncm
        celula_ncm.alignment = Alignment(horizontal="center", vertical="center")
        celula_ncm.font = Font(bold=True, name="Consolas")
        celula_ncm.fill = preenchimento_ncm
        if linha_fim_ncm > linha_inicio_ncm:
            aba.merge_cells(start_row=linha_inicio_ncm, start_column=1, end_row=linha_fim_ncm, end_column=1)

    for indice, largura in enumerate([16, 16, 20, 34], start=1):
        aba.column_dimensions[get_column_letter(indice)].width = largura
    aba.freeze_panes = "A2"


def main() -> None:
    registros = carregar_registros(CAMINHO_JSON)
    diagnosticar_estrutura(registros)

    linhas_referencia = extrair_linhas_referencia(registros)

    workbook = Workbook()
    escrever_aba_plana(workbook, linhas_referencia)

    arvore = construir_arvore(linhas_referencia)
    escrever_aba_arvore(workbook, arvore)

    workbook.save(CAMINHO_SAIDA)
    print(f"Planilha de referência gerada em: {CAMINHO_SAIDA.resolve()}")


if __name__ == "__main__":
    main()