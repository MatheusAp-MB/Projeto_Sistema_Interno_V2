# * [RESUMO] → Gera um relatório Excel comparando o frete "esperado"
#              (dimensão de EMBALAGEM do ERP) com o frete "real"
#              (dimensão declarada pelo vendedor no Mercado Livre) —
#              SÓ LEITURA, não grava nada no banco (diferente de
#              testar_dimensoes_declaradas.py, que também atualiza
#              frete_real). Ferramenta de auditoria/relatório pontual.
#
#              4 abas: Ativos / Pausados / Inativos (pelo status real
#              do anúncio) / "Não foi possível comparar" (quando 1
#              dos 2 lados não tem dado suficiente). Dentro de cada
#              aba de status, 3 seções empilhadas: ML cobra mais /
#              Fretes iguais / ML cobra menos — cada uma ordenada da
#              maior diferença pra menor, com a coluna Diferença
#              colorida (vermelho=cobra mais, verde=cobra menos).

import json
from decimal import Decimal, InvalidOperation
from pathlib import Path
from django.core.management.base import BaseCommand
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

CAMINHO_DETALHES_MLBS = Path('Arquivos_API/detalhes_mlbs.json')
NOME_ARQUIVO_SAIDA = 'Relatorio de comparação de frete ERP vs ML (16-07-2026 16h30).xlsx'

CABECALHOS = [
    'SKU', 'MLB',
    'Altura ERP (cm)', 'Largura ERP (cm)', 'Comprimento ERP (cm)', 'Peso ERP (kg)',
    'Altura ML (cm)', 'Largura ML (cm)', 'Comprimento ML (cm)', 'Peso ML (kg)',
    'Frete ERP (R$)', 'Frete ML (R$)', 'Diferença (R$)',
]

# * [EXPLICAÇÃO] → Estilo padrão do Excel pra "bom"/"ruim" — vermelho
#                  claro quando o ML cobra MAIS (pior pro vendedor),
#                  verde claro quando cobra MENOS.
PREENCHIMENTO_PIOR = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
FONTE_PIOR = Font(color='9C0006')
PREENCHIMENTO_MELHOR = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
FONTE_MELHOR = Font(color='006100')
PREENCHIMENTO_TITULO = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
FONTE_TITULO = Font(color='FFFFFF', bold=True, size=12)
FONTE_CABECALHO = Font(bold=True)
PREENCHIMENTO_CABECALHO = PatternFill(start_color='EEF2F7', end_color='EEF2F7', fill_type='solid')


def _parsear_numero(texto):
    """'19 cm' -> Decimal('19'), '550 g' -> Decimal('550'). None se
    vazio/inválido — nunca derruba o script por 1 registro ruim."""
    if not texto:
        return None
    try:
        primeira_parte = str(texto).strip().split()[0]
        return Decimal(primeira_parte.replace(',', '.'))
    except (InvalidOperation, IndexError, ValueError):
        return None


def _buscar_frete_por_peso_e_preco(peso, preco, frete_todas):
    for faixa in frete_todas:
        peso_ok = faixa.peso_min <= peso and (faixa.peso_max is None or faixa.peso_max >= peso)
        preco_ok = faixa.preco_min <= preco and (faixa.preco_max is None or faixa.preco_max >= preco)
        if peso_ok and preco_ok:
            return faixa.valor
    return None


def _categoria_status(status):
    if status == 'active':
        return 'Ativos'
    if status == 'paused':
        return 'Pausados'
    return 'Inativos'


class Command(BaseCommand):
    help = (
        'Gera um Excel comparando frete "esperado" (ERP, embalagem) com '
        'frete "real" (dimensão declarada no ML) — 4 abas por status + '
        'seções por resultado da comparação. SÓ LEITURA, não grava nada '
        'no banco.'
    )

    def handle(self, *args, **options):
        from mercado_livre.models import AnuncioMercadoLivre, FreteML

        if not CAMINHO_DETALHES_MLBS.exists():
            self.stdout.write(self.style.WARNING(f'Arquivo {CAMINHO_DETALHES_MLBS} não encontrado.'))
            return

        self.stdout.write(f'Lendo {CAMINHO_DETALHES_MLBS}...')
        with open(CAMINHO_DETALHES_MLBS, encoding='utf-8') as f:
            dados = json.load(f)

        registros = dados.get('registros', [])
        self.stdout.write(f'    {len(registros)} registros no JSON')

        anuncios_por_mlb = {
            a.mlb: a for a in AnuncioMercadoLivre.objects.select_related('tipo_de_anuncio')
            .prefetch_related('variacoes__produto').all()
        }
        frete_todas = list(FreteML.objects.all())

        # * [EXPLICAÇÃO] → 3 baldes por status + 1 balde separado pra
        #                  "não foi possível comparar" (qualquer
        #                  status) — populados durante o loop único.
        linhas_por_status = {'Ativos': [], 'Pausados': [], 'Inativos': []}
        linhas_sem_comparacao = []

        sem_anuncio = 0
        sem_variacao = 0
        total = len(registros)

        for indice, reg in enumerate(registros, start=1):
            if indice % 1000 == 0 or indice == total:
                self.stdout.write(f'    ... {indice}/{total} registros processados')

            mlb = reg.get('mlb')
            anuncio = anuncios_por_mlb.get(mlb)
            if not anuncio:
                sem_anuncio += 1
                continue

            status = anuncio.tipo_de_anuncio.status if anuncio.tipo_de_anuncio else 'closed'
            categoria = _categoria_status(status)

            variacao_id = str(reg.get('variacao_id') or mlb)
            variacao = next((v for v in anuncio.variacoes.all() if v.variacao_id == variacao_id), None)
            if not variacao:
                sem_variacao += 1
                continue

            produto = variacao.produto
            sku = produto.sku if produto else '?'

            # --- Lado ML (dimensão declarada) ---
            altura_ml = _parsear_numero(reg.get('attr_seller_package_height'))
            largura_ml = _parsear_numero(reg.get('attr_seller_package_width'))
            comprimento_ml = _parsear_numero(reg.get('attr_seller_package_length'))
            peso_g = _parsear_numero(reg.get('attr_seller_package_weight'))
            peso_ml = (peso_g / 1000) if peso_g is not None else None

            tem_dimensao_ml = altura_ml is not None and largura_ml is not None and comprimento_ml is not None

            if not tem_dimensao_ml and peso_ml is None:
                peso_legado_g = _parsear_numero(reg.get('attr_weight'))
                if peso_legado_g is not None:
                    peso_ml = peso_legado_g / 1000

            peso_volumetrico_ml = None
            if tem_dimensao_ml:
                peso_volumetrico_ml = (altura_ml * largura_ml * comprimento_ml) / Decimal('6000')

            candidatos_peso_ml = [p for p in (peso_ml, peso_volumetrico_ml) if p is not None]
            peso_real_ml = max(candidatos_peso_ml) if candidatos_peso_ml else None

            frete_ml = None
            if peso_real_ml is not None and variacao.preco_atual:
                frete_ml = _buscar_frete_por_peso_e_preco(peso_real_ml, variacao.preco_atual, frete_todas)

            # --- Lado ERP (embalagem do produto) ---
            peso_erp = None
            frete_erp = None
            if produto:
                peso_cubado = produto.peso_cubado or Decimal('0')
                peso_embalagem = produto.peso_produto_apos_embalado or Decimal('0')
                peso_erp = max(peso_embalagem, peso_cubado)
                if peso_erp > 0 and variacao.preco_atual:
                    frete_erp = _buscar_frete_por_peso_e_preco(peso_erp, variacao.preco_atual, frete_todas)

            linha_base = {
                'sku': sku,
                'mlb': mlb,
                'altura_erp': produto.altura_produto_apos_embalado if produto else None,
                'largura_erp': produto.largura_produto_apos_embalado if produto else None,
                'comprimento_erp': produto.comprimento_produto_apos_embalado if produto else None,
                'peso_erp': peso_erp,
                'altura_ml': altura_ml,
                'largura_ml': largura_ml,
                'comprimento_ml': comprimento_ml,
                'peso_ml': peso_real_ml,
                'frete_erp': frete_erp,
                'frete_ml': frete_ml,
            }

            # * [EXPLICAÇÃO] → Comparação só é justa com dado real dos
            #                  2 lados (mesmo critério já validado
            #                  antes) — senão vai pra aba separada.
            tem_dado_erp = peso_erp is not None and peso_erp > 0 and frete_erp is not None
            tem_dado_ml = tem_dimensao_ml and frete_ml is not None

            if not tem_dado_erp or not tem_dado_ml:
                linha_base['status'] = categoria
                linha_base['motivo'] = (
                    'Sem embalagem cadastrada no ERP' if not tem_dado_erp
                    else 'Sem dimensão declarada no ML (só peso, ou nenhum dado)'
                )
                linhas_sem_comparacao.append(linha_base)
                continue

            linha_base['diferenca'] = frete_ml - frete_erp
            linhas_por_status[categoria].append(linha_base)

        self.stdout.write(self.style.SUCCESS(
            f'Processamento concluído. Sem anúncio: {sem_anuncio}, sem variação: {sem_variacao}.'
        ))
        self.stdout.write('Montando o Excel...')

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for categoria in ('Ativos', 'Pausados', 'Inativos'):
            self._montar_aba_status(wb, categoria, linhas_por_status[categoria])

        self._montar_aba_sem_comparacao(wb, linhas_sem_comparacao)

        wb.save(NOME_ARQUIVO_SAIDA)
        self.stdout.write(self.style.SUCCESS(f'Relatório salvo em: {NOME_ARQUIVO_SAIDA}'))

    def _escrever_cabecalho_secao(self, ws, linha, titulo, largura_colunas):
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=largura_colunas)
        celula = ws.cell(row=linha, column=1, value=titulo)
        celula.fill = PREENCHIMENTO_TITULO
        celula.font = FONTE_TITULO
        celula.alignment = Alignment(horizontal='center')
        return linha + 1

    def _escrever_cabecalho_tabela(self, ws, linha):
        for coluna, texto in enumerate(CABECALHOS, start=1):
            celula = ws.cell(row=linha, column=coluna, value=texto)
            celula.font = FONTE_CABECALHO
            celula.fill = PREENCHIMENTO_CABECALHO
        return linha + 1

    def _escrever_linhas(self, ws, linha, lista_linhas, preenchimento_diferenca=None, fonte_diferenca=None):
        for item in lista_linhas:
            valores = [
                item['sku'], item['mlb'],
                item['altura_erp'], item['largura_erp'], item['comprimento_erp'], item['peso_erp'],
                item['altura_ml'], item['largura_ml'], item['comprimento_ml'], item['peso_ml'],
                item['frete_erp'], item['frete_ml'], item['diferenca'],
            ]
            for coluna, valor in enumerate(valores, start=1):
                celula = ws.cell(row=linha, column=coluna, value=float(valor) if isinstance(valor, Decimal) else valor)
                if coluna == 13 and preenchimento_diferenca is not None:
                    celula.fill = preenchimento_diferenca
                    celula.font = fonte_diferenca
            linha += 1
        return linha

    def _montar_aba_status(self, wb, nome_aba, linhas):
        ws = wb.create_sheet(nome_aba)

        maiores = sorted([l for l in linhas if l['diferenca'] > 0], key=lambda l: l['diferenca'], reverse=True)
        iguais = sorted([l for l in linhas if l['diferenca'] == 0], key=lambda l: l['sku'])
        menores = sorted([l for l in linhas if l['diferenca'] < 0], key=lambda l: l['diferenca'])

        linha = 1
        linha = self._escrever_cabecalho_secao(ws, linha, f'Mercado Livre cobrando MAIS ({len(maiores)})', len(CABECALHOS))
        linha = self._escrever_cabecalho_tabela(ws, linha)
        linha = self._escrever_linhas(ws, linha, maiores, PREENCHIMENTO_PIOR, FONTE_PIOR)
        linha += 1

        linha = self._escrever_cabecalho_secao(ws, linha, f'Fretes iguais ({len(iguais)})', len(CABECALHOS))
        linha = self._escrever_cabecalho_tabela(ws, linha)
        linha = self._escrever_linhas(ws, linha, iguais)
        linha += 1

        linha = self._escrever_cabecalho_secao(ws, linha, f'Mercado Livre cobrando MENOS ({len(menores)})', len(CABECALHOS))
        linha = self._escrever_cabecalho_tabela(ws, linha)
        linha = self._escrever_linhas(ws, linha, menores, PREENCHIMENTO_MELHOR, FONTE_MELHOR)

        for coluna in range(1, len(CABECALHOS) + 1):
            ws.column_dimensions[get_column_letter(coluna)].width = 16

    def _montar_aba_sem_comparacao(self, wb, linhas):
        ws = wb.create_sheet('Não foi possível comparar')

        cabecalhos = ['SKU', 'MLB', 'Status',
                      'Altura ERP (cm)', 'Largura ERP (cm)', 'Comprimento ERP (cm)', 'Peso ERP (kg)',
                      'Altura ML (cm)', 'Largura ML (cm)', 'Comprimento ML (cm)', 'Peso ML (kg)',
                      'Frete ERP (R$)', 'Frete ML (R$)', 'Motivo']

        for coluna, texto in enumerate(cabecalhos, start=1):
            celula = ws.cell(row=1, column=coluna, value=texto)
            celula.font = FONTE_CABECALHO
            celula.fill = PREENCHIMENTO_CABECALHO

        linha = 2
        for item in sorted(linhas, key=lambda l: (l['status'], l['sku'])):
            valores = [
                item['sku'], item['mlb'], item['status'],
                item['altura_erp'], item['largura_erp'], item['comprimento_erp'], item['peso_erp'],
                item['altura_ml'], item['largura_ml'], item['comprimento_ml'], item['peso_ml'],
                item['frete_erp'], item['frete_ml'], item['motivo'],
            ]
            for coluna, valor in enumerate(valores, start=1):
                ws.cell(row=linha, column=coluna, value=float(valor) if isinstance(valor, Decimal) else valor)
            linha += 1

        for coluna in range(1, len(cabecalhos) + 1):
            ws.column_dimensions[get_column_letter(coluna)].width = 16