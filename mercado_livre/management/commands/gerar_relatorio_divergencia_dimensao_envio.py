# mercado_livre/management/commands/gerar_relatorio_divergencia_dimensao_envio.py

# Função Objetivo: Gera um Excel detalhado de todos os MLBs com situação de dimensão de
# envio persistida — pra auditoria e handoff (correção manual via API do ML).
# Explicação em detalhe: SÓ LEITURA, não grava nada no banco. 6 abas: Resumo (contagem por
# situação + pelas 3 categorias de causa dentro de Divergente) + 1 aba por categoria de causa
# (ML usa Sem Embalar / Offset uniforme outro padrão / Sem padrão) + 1 aba com os outros 5
# estados (Iguais, Não Refletida ML, Não Salva ERP, Sem Dado, Sem Produto), pra referência
# completa num só arquivo.

import json
from decimal import Decimal
from pathlib import Path
from django.core.management.base import BaseCommand
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from mercado_livre.models import VariacaoAnuncioMercadoLivre
from mercado_livre.funcoes_auxiliares.comparador_dimensao_envio import SituacaoDimensaoEnvio

NOME_ARQUIVO_SAIDA = 'Relatorio de Divergencia de Dimensao de Envio (ERP vs ML).xlsx'
NOME_ARQUIVO_SAIDA_JSON = 'Relatorio de Divergencia de Dimensao de Envio (ERP vs ML).json'

CABECALHOS_DETALHE = [
    'SKU', 'EAN', 'Título', 'Marca', 'MLB', 'Permalink',
    'Tipo de Anúncio', 'Status', 'Classificação Catálogo',
    'ERP c/ Embalagem — Altura', 'ERP c/ Embalagem — Largura', 'ERP c/ Embalagem — Comprimento',
    'ERP c/ Embalagem — Peso',
    'ERP Sem Embalar — Altura', 'ERP Sem Embalar — Largura', 'ERP Sem Embalar — Comprimento',
    'ERP Sem Embalar — Peso',
    'ML Declarado — Altura', 'ML Declarado — Largura', 'ML Declarado — Comprimento',
    'ML Declarado — Peso',
    'Maior Diferença Dimensão (cm)', 'Diferença Peso (kg)',
]

PREENCHIMENTO_TITULO = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
FONTE_TITULO = Font(color='FFFFFF', bold=True, size=12)
FONTE_CABECALHO = Font(bold=True)
PREENCHIMENTO_CABECALHO = PatternFill(start_color='EEF2F7', end_color='EEF2F7', fill_type='solid')
PREENCHIMENTO_ALERTA = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
FONTE_ALERTA = Font(color='9C0006')


# Função Objetivo: Classifica 1 variação divergente numa das 3 categorias de causa.
def classificar_causa_divergencia(variacao, produto):
    dims_erp_com = [produto.altura_ordenada_cm, produto.largura_ordenada_cm, produto.comprimento_ordenada_cm]
    dims_ml = [variacao.altura_ordenada_cm, variacao.largura_ordenada_cm, variacao.comprimento_ordenada_cm]

    dims_erp_sem_brutas = [
        produto.altura_produto_sem_embalar, produto.largura_produto_sem_embalar,
        produto.comprimento_produto_sem_embalar,
    ]
    sem_embalar_tem_dado = any(valor != 0 for valor in dims_erp_sem_brutas) or produto.peso_produto_sem_embalar != 0
    dims_erp_sem_ordenadas = sorted(dims_erp_sem_brutas)

    bate_com_sem_embalar = (
        sem_embalar_tem_dado
        and dims_erp_sem_ordenadas == dims_ml
        and produto.peso_produto_sem_embalar == variacao.peso_declarado_kg
    )
    if bate_com_sem_embalar:
        return 'ML usa SEM EMBALAR do ERP (confirmado)'

    diferencas = [erp - ml for erp, ml in zip(dims_erp_com, dims_ml)]
    offset_uniforme = len(set(diferencas)) == 1 and diferencas[0] != 0
    if offset_uniforme:
        return 'Offset uniforme — outro padrão'

    return 'Sem padrão — divergência real'


# Função Objetivo: Monta 1 linha de detalhe (dict) a partir de 1 variação + seu produto.
def montar_linha_detalhe(variacao, produto):
    dims_erp_com = [produto.altura_ordenada_cm, produto.largura_ordenada_cm, produto.comprimento_ordenada_cm]
    dims_ml = [variacao.altura_ordenada_cm, variacao.largura_ordenada_cm, variacao.comprimento_ordenada_cm]

    diferencas_validas = [abs(erp - ml) for erp, ml in zip(dims_erp_com, dims_ml) if erp is not None and ml is not None]
    maior_diferenca = max(diferencas_validas, default=None)

    diferenca_peso = None
    if produto.peso_produto_apos_embalado is not None and variacao.peso_declarado_kg is not None:
        diferenca_peso = produto.peso_produto_apos_embalado - variacao.peso_declarado_kg

    tipo_anuncio = variacao.anuncio.tipo_de_anuncio

    return {
        'sku': produto.sku,
        'ean': produto.ean,
        'titulo': produto.titulo,
        'marca': produto.marca,
        'mlb': variacao.anuncio.mlb,
        'permalink': variacao.anuncio.permalink,
        'tipo_anuncio': tipo_anuncio.get_tipo_anuncio_display() if tipo_anuncio else None,
        'status': tipo_anuncio.get_status_display() if tipo_anuncio else None,
        'classificacao_catalogo': tipo_anuncio.get_classificacao_catalogo_display() if tipo_anuncio else None,
        'erp_com_altura': produto.altura_ordenada_cm,
        'erp_com_largura': produto.largura_ordenada_cm,
        'erp_com_comprimento': produto.comprimento_ordenada_cm,
        'erp_com_peso': produto.peso_produto_apos_embalado,
        'erp_sem_altura': produto.altura_produto_sem_embalar,
        'erp_sem_largura': produto.largura_produto_sem_embalar,
        'erp_sem_comprimento': produto.comprimento_produto_sem_embalar,
        'erp_sem_peso': produto.peso_produto_sem_embalar,
        'ml_altura': variacao.altura_ordenada_cm,
        'ml_largura': variacao.largura_ordenada_cm,
        'ml_comprimento': variacao.comprimento_ordenada_cm,
        'ml_peso': variacao.peso_declarado_kg,
        'maior_diferenca': maior_diferenca,
        'diferenca_peso': diferenca_peso,
    }


# Função Objetivo: Converte Decimal pra float — JSON nativo não serializa Decimal.
def _serializar_valor_json(valor):
    if isinstance(valor, Decimal):
        return float(valor)
    return valor


# Função Objetivo: Serializa 1 linha de detalhe (dict com Decimals) pra tipos nativos de JSON.
def _serializar_linha_json(linha):
    return {chave: _serializar_valor_json(valor) for chave, valor in linha.items()}


# Função Objetivo: Gera o JSON a partir dos MESMOS dados já categorizados usados no Excel.
# Explicação em detalhe: pensado pra facilitar a leitura por LLM (colar direto no chat,
# estrutura sem ambiguidade) — o Excel continua existindo em paralelo pra revisão visual.
def gerar_json_divergencia_dimensao_envio(total, linhas_por_categoria, linhas_outras_situacoes,
                                            caminho_saida=NOME_ARQUIVO_SAIDA_JSON):
    dados = {
        'total_variacoes_com_situacao_calculada': total,
        'categorias_divergencia': {
            categoria: [_serializar_linha_json(linha) for linha in linhas]
            for categoria, linhas in linhas_por_categoria.items()
        },
        'outras_situacoes': {
            label: [_serializar_linha_json(linha) for linha in linhas]
            for label, linhas in linhas_outras_situacoes.items()
        },
    }
    with open(caminho_saida, 'w', encoding='utf-8') as arquivo:
        json.dump(dados, arquivo, ensure_ascii=False, indent=2)
    return caminho_saida


class Command(BaseCommand):
    help = (
        'Gera um Excel + JSON detalhados de todos os MLBs com situação de dimensão de envio '
        'persistida, categorizando a causa da divergência. SÓ LEITURA, não grava nada.'
    )

    def handle(self, *args, **options):
        variacoes = list(
            VariacaoAnuncioMercadoLivre.objects
            .filter(situacao_dimensao_envio__isnull=False)
            .select_related('produto', 'anuncio', 'anuncio__tipo_de_anuncio')
        )
        self.stdout.write(f'{len(variacoes)} variações com situação de dimensão de envio persistida.')

        linhas_por_categoria = {
            'ML usa SEM EMBALAR do ERP (confirmado)': [],
            'Offset uniforme — outro padrão': [],
            'Sem padrão — divergência real': [],
        }
        linhas_outras_situacoes = {estado.label: [] for estado in SituacaoDimensaoEnvio if estado != SituacaoDimensaoEnvio.DIVERGENTE}

        for variacao in variacoes:
            situacao = SituacaoDimensaoEnvio(variacao.situacao_dimensao_envio)

            if situacao != SituacaoDimensaoEnvio.DIVERGENTE:
                if variacao.produto is not None:
                    linha = montar_linha_detalhe(variacao, variacao.produto)
                else:
                    linha = {'sku': None, 'ean': None, 'titulo': None, 'marca': None,
                              'mlb': variacao.anuncio.mlb, 'permalink': variacao.anuncio.permalink,
                              'tipo_anuncio': None, 'status': None, 'classificacao_catalogo': None,
                              'erp_com_altura': None, 'erp_com_largura': None, 'erp_com_comprimento': None,
                              'erp_com_peso': None, 'erp_sem_altura': None, 'erp_sem_largura': None,
                              'erp_sem_comprimento': None, 'erp_sem_peso': None,
                              'ml_altura': variacao.altura_ordenada_cm, 'ml_largura': variacao.largura_ordenada_cm,
                              'ml_comprimento': variacao.comprimento_ordenada_cm, 'ml_peso': variacao.peso_declarado_kg,
                              'maior_diferenca': None, 'diferenca_peso': None}
                linhas_outras_situacoes[situacao.label].append(linha)
                continue

            produto = variacao.produto
            categoria = classificar_causa_divergencia(variacao, produto)
            linha = montar_linha_detalhe(variacao, produto)
            linhas_por_categoria[categoria].append(linha)

        for categoria in linhas_por_categoria:
            linhas_por_categoria[categoria].sort(key=lambda linha: linha['maior_diferenca'] or 0, reverse=True)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        self._montar_aba_resumo(wb, len(variacoes), linhas_por_categoria, linhas_outras_situacoes)
        for categoria, linhas in linhas_por_categoria.items():
            self._montar_aba_detalhe(wb, categoria, linhas, destacar=(categoria != 'Sem padrão — divergência real'))
        for label, linhas in linhas_outras_situacoes.items():
            if linhas:
                self._montar_aba_detalhe(wb, label, linhas, destacar=False)

        wb.save(NOME_ARQUIVO_SAIDA)
        self.stdout.write(self.style.SUCCESS(f'Relatório (Excel) salvo em: {NOME_ARQUIVO_SAIDA}'))

        caminho_json = gerar_json_divergencia_dimensao_envio(len(variacoes), linhas_por_categoria, linhas_outras_situacoes)
        self.stdout.write(self.style.SUCCESS(f'Relatório (JSON) salvo em: {caminho_json}'))

    def _montar_aba_resumo(self, wb, total, linhas_por_categoria, linhas_outras_situacoes):
        ws = wb.create_sheet('Resumo')
        linha = 1

        ws.cell(row=linha, column=1, value='RESUMO — DIVERGÊNCIA DE DIMENSÃO DE ENVIO (ERP vs ML)').font = Font(bold=True, size=14)
        linha += 2

        ws.cell(row=linha, column=1, value=f'Total de variações com situação calculada: {total}').font = Font(bold=True)
        linha += 2

        ws.cell(row=linha, column=1, value='Categorias dentro de DIVERGENTE (por prioridade de correção):').font = FONTE_CABECALHO
        linha += 1
        for categoria, linhas_categoria in linhas_por_categoria.items():
            ws.cell(row=linha, column=1, value=f'  {categoria}: {len(linhas_categoria)}')
            linha += 1
        linha += 1

        ws.cell(row=linha, column=1, value='Outras situações (fora de Divergente):').font = FONTE_CABECALHO
        linha += 1
        for label, linhas_situacao in linhas_outras_situacoes.items():
            ws.cell(row=linha, column=1, value=f'  {label}: {len(linhas_situacao)}')
            linha += 1

        ws.column_dimensions['A'].width = 70

    def _montar_aba_detalhe(self, wb, nome_categoria, linhas, destacar):
        nome_aba = nome_categoria[:31]
        ws = wb.create_sheet(nome_aba)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(CABECALHOS_DETALHE))
        celula_titulo = ws.cell(row=1, column=1, value=f'{nome_categoria} ({len(linhas)})')
        celula_titulo.fill = PREENCHIMENTO_TITULO
        celula_titulo.font = FONTE_TITULO
        celula_titulo.alignment = Alignment(horizontal='center')

        for coluna, texto in enumerate(CABECALHOS_DETALHE, start=1):
            celula = ws.cell(row=2, column=coluna, value=texto)
            celula.font = FONTE_CABECALHO
            celula.fill = PREENCHIMENTO_CABECALHO

        linha_atual = 3
        for item in linhas:
            valores = [
                item['sku'], item['ean'], item['titulo'], item['marca'], item['mlb'], item['permalink'],
                item['tipo_anuncio'], item['status'], item['classificacao_catalogo'],
                item['erp_com_altura'], item['erp_com_largura'], item['erp_com_comprimento'], item['erp_com_peso'],
                item['erp_sem_altura'], item['erp_sem_largura'], item['erp_sem_comprimento'], item['erp_sem_peso'],
                item['ml_altura'], item['ml_largura'], item['ml_comprimento'], item['ml_peso'],
                item['maior_diferenca'], item['diferenca_peso'],
            ]
            for coluna, valor in enumerate(valores, start=1):
                celula = ws.cell(row=linha_atual, column=coluna, value=float(valor) if hasattr(valor, 'as_tuple') else valor)
                if destacar and coluna == 22 and valor is not None and valor > 10:
                    celula.fill = PREENCHIMENTO_ALERTA
                    celula.font = FONTE_ALERTA
            linha_atual += 1

        for coluna in range(1, len(CABECALHOS_DETALHE) + 1):
            ws.column_dimensions[get_column_letter(coluna)].width = 18