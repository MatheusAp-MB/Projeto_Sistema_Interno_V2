# mercado_livre/funcoes_auxiliares/exportacao_resumo_criterios.py

# Função Objetivo: Gera o .xlsx do Resumo de Critérios, respeitando os mesmos filtros
# da tela — mesmo texto exibido no HTML (SIM/NÃO/N/A/Não calculado), sem paginação
# (traz tudo que bate com o filtro).

import io
from dataclasses import dataclass
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from mercado_livre.funcoes_auxiliares.badges import (
    BADGES_STATUS, BADGES_TIPO_ANUNCIO, BADGES_LOGISTICA, BADGES_CATALOGO, badge_de,
)

# * [EXPLICAÇÃO] → Mesmas cores exatas usadas na tela (celula-sim/nao/na/
#                  nao-calculado, layout_resumo_criterios.css) — o Excel
#                  precisa ter o mesmo significado visual que o usuário
#                  já reconhece na tela, não uma paleta nova inventada aqui.
FONTE_PADRAO = Font(name='Arial', size=10)
FONTE_CABECALHO = Font(name='Arial', size=10, bold=True, color='FFFFFF')

PREENCHIMENTO_CABECALHO_FIXO = PatternFill('solid', fgColor='1E3A5F')

ESTILO_POR_TEXTO_CRITERIO = {
    'SIM':           {'fill': 'D4EDDA', 'fonte': Font(name='Arial', size=10, bold=True, color='1D6B3A')},
    'NÃO':           {'fill': 'F8D7DA', 'fonte': Font(name='Arial', size=10, bold=True, color='A31C2F')},
    'N/A':           {'fill': 'F1F2F4', 'fonte': Font(name='Arial', size=10, bold=True, color='8A8F98')},
    'Não calculado': {'fill': 'E9E9E9', 'fonte': Font(name='Arial', size=10, italic=True, color='999999')},
}

TEXTO_POR_STATUS_CRITERIO = {
    'aprovado': 'SIM',
    'nao_aprovado': 'NÃO',
    'nao_aplicavel': 'N/A',
}


@dataclass
class ResultadoCriterioExportado:
    texto: str
    link_correcao: str = None


@dataclass
class LinhaExportacaoResumoCriterios:
    sku: str
    mlb: str
    marca: str
    titulo: str
    status: str
    tipo_anuncio: str
    logistica: str
    flex: str
    catalogo: str
    score: object
    nivel: str
    resultados_criterios: list  # lista de ResultadoCriterioExportado, na ordem de `criterios`


# Função Objetivo: Converte as variações filtradas em linhas prontas pra exportar —
# mesma lógica de exibição da tela (view_resumo_criterios), sem paginação.
def montar_linhas_exportacao(variacoes, criterios):
    linhas = []

    for variacao in variacoes:
        anuncio = variacao.anuncio
        tipo = anuncio.tipo_de_anuncio
        qualidade = getattr(variacao, 'qualidade', None)

        if qualidade is None:
            resultados_criterios = [ResultadoCriterioExportado('Não calculado') for _ in criterios]
            score, nivel = None, None
        else:
            avaliacao_por_rule_key = {
                avaliacao.criterio.rule_key: avaliacao
                for avaliacao in qualidade.criterios.all()
            }
            resultados_criterios = []
            for c in criterios:
                avaliacao = avaliacao_por_rule_key.get(c.rule_key)
                texto = TEXTO_POR_STATUS_CRITERIO.get(avaliacao.status if avaliacao else None, 'Não calculado')
                link = avaliacao.link_correcao if (avaliacao and avaliacao.status == 'nao_aprovado') else None
                resultados_criterios.append(ResultadoCriterioExportado(texto, link))
            score, nivel = qualidade.score, qualidade.nivel

        linhas.append(LinhaExportacaoResumoCriterios(
            sku=variacao.produto.sku,
            mlb=anuncio.mlb,
            marca=variacao.produto.marca,
            titulo=anuncio.titulo_anuncio,
            status=badge_de(BADGES_STATUS, tipo.status)['label'] if tipo else '—',
            tipo_anuncio=badge_de(BADGES_TIPO_ANUNCIO, tipo.tipo_anuncio)['label'] if tipo else '—',
            logistica=badge_de(BADGES_LOGISTICA, tipo.tipo_logistico)['label'] if tipo else '—',
            flex=('Com Flex' if (tipo and tipo.flex) else 'Sem Flex') if tipo else '—',
            catalogo=badge_de(BADGES_CATALOGO, tipo.classificacao_catalogo)['label'] if tipo else '—',
            score=score, nivel=nivel,
            resultados_criterios=resultados_criterios,
        ))

    return linhas


COLUNAS_FIXAS = [
    'SKU', 'MLB', 'Marca', 'Título', 'Situação', 'Tipo de Anúncio',
    'Logística', 'Flex', 'Situação do Catálogo', 'Score', 'Nível',
]


def gerar_excel_resumo_criterios(linhas, criterios):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Resumo de Critérios'

    cabecalho = COLUNAS_FIXAS + [c.pergunta for c in criterios]
    ws.append(cabecalho)

    for indice, celula in enumerate(ws[1], start=1):
        celula.font = FONTE_CABECALHO
        celula.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if indice <= len(COLUNAS_FIXAS):
            celula.fill = PREENCHIMENTO_CABECALHO_FIXO
        else:
            cor_criterio = criterios[indice - len(COLUNAS_FIXAS) - 1].cor_resolvida
            celula.fill = PatternFill('solid', fgColor=cor_criterio.lstrip('#').upper())

    for linha in linhas:
        valores = [
            linha.sku, linha.mlb, linha.marca, linha.titulo, linha.status,
            linha.tipo_anuncio, linha.logistica, linha.flex, linha.catalogo,
            linha.score if linha.score is not None else '—',
            linha.nivel if linha.nivel else '—',
            *[resultado.texto for resultado in linha.resultados_criterios],
        ]
        ws.append(valores)

        linha_excel = ws.max_row
        for indice_coluna, resultado in enumerate(linha.resultados_criterios, start=len(COLUNAS_FIXAS) + 1):
            estilo = ESTILO_POR_TEXTO_CRITERIO.get(resultado.texto)
            celula = ws.cell(row=linha_excel, column=indice_coluna)
            if estilo:
                celula.fill = PatternFill('solid', fgColor=estilo['fill'])
                celula.font = estilo['fonte']
            else:
                celula.font = FONTE_PADRAO
            celula.alignment = Alignment(horizontal='center')

            # * [EXPLICAÇÃO] → Link de correção vira hyperlink de verdade na
            #                  célula (clicável no Excel), mantendo o texto
            #                  "NÃO" visível — sublinhado avisa que é clicável.
            if resultado.link_correcao:
                celula.hyperlink = resultado.link_correcao
                celula.font = Font(
                    name='Arial', size=10, bold=True, color=estilo['fonte'].color, underline='single',
                )

        for indice_coluna in range(1, len(COLUNAS_FIXAS) + 1):
            ws.cell(row=linha_excel, column=indice_coluna).font = FONTE_PADRAO

    # * [EXPLICAÇÃO] → Congela cabeçalho + as 4 primeiras colunas (SKU/MLB/
    #                  Marca/Título) — rolando a planilha pra direita ou
    #                  pra baixo, essas referências continuam visíveis.
    ws.freeze_panes = 'E2'
    ws.auto_filter.ref = ws.dimensions

    LARGURA_POR_COLUNA_FIXA = {
        'SKU': 16, 'MLB': 14, 'Marca': 18, 'Título': 48, 'Situação': 14,
        'Tipo de Anúncio': 14, 'Logística': 12, 'Flex': 10,
        'Situação do Catálogo': 18, 'Score': 8, 'Nível': 10,
    }
    for indice, nome in enumerate(COLUNAS_FIXAS, start=1):
        ws.column_dimensions[get_column_letter(indice)].width = LARGURA_POR_COLUNA_FIXA[nome]
    for indice in range(len(COLUNAS_FIXAS) + 1, len(cabecalho) + 1):
        ws.column_dimensions[get_column_letter(indice)].width = 22

    ws.row_dimensions[1].height = 34

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()