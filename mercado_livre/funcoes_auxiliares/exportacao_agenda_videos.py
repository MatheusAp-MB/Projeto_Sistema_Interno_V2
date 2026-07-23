# mercado_livre/funcoes_auxiliares/exportacao_agenda_videos.py

# Função Objetivo: Gera o .xlsx da "Agenda de Vídeos" — 1 linha por produto único,
# no formato que alimenta a planilha semi-automática externa de agendamento.
# Explicação em detalhe: reaproveita o MESMO filtro já aplicado na tela de Resumo
# de Critérios (sem nenhum filtro forçado no servidor — confia no que o usuário
# escolheu, ex: Status=Ativo + critério "Sem clipes"=Não aprovado).

import io
from dataclasses import dataclass
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

URL_BASE_CLIPES = 'https://www.mercadolivre.com.br/video/creator/?item_id={mlb}#from=menu_mkt'

# * [EXPLICAÇÃO] → Fixo, independente de quais critérios o usuário estiver
#                  filtrando na tela ao mesmo tempo — a Agenda de Vídeos
#                  sempre busca o link de correção deste critério específico.
RULE_KEY_VIDEO = 'UP_HAS_SHORTS'

CABECALHO = [
    'Empresa', 'Código de Barras', 'Título do Produto', 'Marca', 'Data', 'Status',
    'Link de Correção (Vídeo)', 'Link para a tela de clipes',
]


@dataclass
class LinhaAgendaVideos:
    ean: str
    titulo: str
    marca: str
    link_correcao: str  # pode ser None — produto sem avaliação de UP_HAS_SHORTS
    link_clipes: str


# Função Objetivo: Deduplica por produto — 1 MLB por SKU (a 1ª variação
# encontrada que já bateu no filtro da tela; as demais do mesmo produto são
# ignoradas). Busca o link_correcao especificamente do critério UP_HAS_SHORTS,
# independente de quais critérios estiverem sendo usados pra filtrar a lista.
def montar_linhas_agenda_videos(variacoes):
    produtos_ja_incluidos = set()
    linhas = []

    for variacao in variacoes:
        produto = variacao.produto
        if produto.id in produtos_ja_incluidos:
            continue
        produtos_ja_incluidos.add(produto.id)

        anuncio = variacao.anuncio
        qualidade = getattr(variacao, 'qualidade', None)

        link_correcao = None
        if qualidade is not None:
            avaliacao_video = next(
                (a for a in qualidade.criterios.all() if a.criterio.rule_key == RULE_KEY_VIDEO), None,
            )
            if avaliacao_video is not None:
                link_correcao = avaliacao_video.link_correcao

        linhas.append(LinhaAgendaVideos(
            ean=produto.ean,
            titulo=produto.titulo,
            marca=produto.marca,
            link_correcao=link_correcao,
            link_clipes=URL_BASE_CLIPES.format(mlb=anuncio.mlb),
        ))

    return linhas


def gerar_excel_agenda_videos(linhas):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Agenda'

    ws.append(CABECALHO)
    for celula in ws[1]:
        celula.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        celula.fill = PatternFill('solid', fgColor='1E3A5F')
        celula.alignment = Alignment(horizontal='center', vertical='center')

    for linha in linhas:
        ws.append([
            'MAGAZINE', linha.ean, linha.titulo, linha.marca, '', 'Ativo',
            linha.link_correcao or '', linha.link_clipes,
        ])
        linha_excel = ws.max_row

        if linha.link_correcao:
            celula_correcao = ws.cell(row=linha_excel, column=7)
            celula_correcao.hyperlink = linha.link_correcao
            celula_correcao.font = Font(name='Arial', size=10, underline='single', color='1155CC')

        celula_clipes = ws.cell(row=linha_excel, column=8)
        celula_clipes.hyperlink = linha.link_clipes
        celula_clipes.font = Font(name='Arial', size=10, underline='single', color='1155CC')

    for coluna, largura in zip(range(1, 9), [12, 18, 48, 18, 10, 10, 42, 55]):
        ws.column_dimensions[get_column_letter(coluna)].width = largura

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 22

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()