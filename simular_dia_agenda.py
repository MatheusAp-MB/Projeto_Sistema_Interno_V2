import re
import sys
import time
import threading
import tkinter as tk
from tkinter import messagebox
from pathlib import Path

import keyboard
import win32gui
import openpyxl
from pywinauto import Application
from pywinauto.keyboard import send_keys

from agente_local.aviso_execucao import AvisoExecucao

# ==== CONFIGURA AQUI ANTES DE RODAR ====
CAMINHO_PLANILHA = Path(r"C:\Users\Win10\Desktop\Videos ML\MAGAZINE.XLSX")
PASTA_RAIZ = Path(r"C:\Users\Win10\Desktop\Videos ML")
NOME_EMPRESA_FIXO = "Magazine"

TECLA_CAPTURA = "F8"
NOME_BOTAO_UPLOAD = "Subir archivo"
NOME_BOTAO_ENVIAR = "Anunciar"

STATUS_FILTRO_ENTRADA = "A fazer hoje"
STATUS_SAIDA = "Aguardando aprovação do ML"
PREFIXO_USADO = "usado - "
# ========================================


# Função Objetivo: Remove caracteres inválidos pra nome de pasta/arquivo no Windows.
def sanitizar_nome(texto):
    texto = re.sub(r'[\\/:*?"<>|]', '', texto)
    return texto.strip(" .")


# Função Objetivo: Acha o Roteiro de menor número ainda disponível (sem o prefixo "usado - ").
def achar_proximo_roteiro(pasta_produto):
    candidatos = []
    if not pasta_produto.exists():
        return None
    for arquivo in pasta_produto.glob("Roteiro *.mp4"):
        m = re.match(r"Roteiro (\d+)\.mp4$", arquivo.name)
        if m:
            candidatos.append((int(m.group(1)), arquivo))
    if not candidatos:
        return None
    candidatos.sort(key=lambda x: x[0])
    return candidatos[0][1]


def montar_url(mlb):
    return f"https://www.mercadolivre.com.br/video/creator/upload?item_syi={mlb}&item_id={mlb}&origin=syi"


# ---------- Captura da janela do Chrome (F8) ----------
janela_capturada = {"handle": None}
evento_capturado = threading.Event()


def capturar_janela_em_primeiro_plano():
    janela_capturada["handle"] = win32gui.GetForegroundWindow()
    evento_capturado.set()


def listar_janelas_abertas():
    handles = []
    win32gui.EnumWindows(lambda h, resultado: resultado.append(h), handles)
    return set(h for h in handles if win32gui.IsWindowVisible(h) and win32gui.GetWindowText(h))


keyboard.add_hotkey(TECLA_CAPTURA, capturar_janela_em_primeiro_plano)

root = tk.Tk()
root.title("Aguardando captura")
root.attributes("-topmost", True)
root.geometry("460x140")
tk.Label(
    root, justify="left", padx=15, pady=15,
    text=(
        "1) Clique na aba do Chrome, já logada no Mercado Livre.\n\n"
        f"2) Sem clicar em mais nada, aperte {TECLA_CAPTURA}."
    ),
).pack()


def verificar_captura():
    if evento_capturado.is_set():
        root.destroy()
    else:
        root.after(200, verificar_captura)


root.after(200, verificar_captura)
root.mainloop()
keyboard.remove_hotkey(TECLA_CAPTURA)

if janela_capturada["handle"] is None:
    print("Cancelado.")
    sys.exit(0)

aviso = AvisoExecucao()
aviso.atualizar("SCRIPT EM EXECUÇÃO — NÃO MEXA NO MOUSE/TECLADO", "#c0392b")

app = Application(backend="uia").connect(handle=janela_capturada["handle"])
janela_ml = app.window(handle=janela_capturada["handle"])

# ---------- Lê a planilha ----------
wb = openpyxl.load_workbook(CAMINHO_PLANILHA, data_only=False)
ws = wb["Agenda"]

linhas_a_processar = []
for numero_linha in range(2, ws.max_row + 1):
    ean = ws.cell(row=numero_linha, column=3).value
    mlb = ws.cell(row=numero_linha, column=4).value
    produto = ws.cell(row=numero_linha, column=5).value
    marca = ws.cell(row=numero_linha, column=6).value
    status = ws.cell(row=numero_linha, column=9).value

    if status == STATUS_FILTRO_ENTRADA and produto and ean:
        linhas_a_processar.append((numero_linha, ean, mlb, produto, marca))

print(f"{len(linhas_a_processar)} produto(s) com status '{STATUS_FILTRO_ENTRADA}'.")

resultados = []

for indice, (numero_linha, ean, mlb, produto, marca) in enumerate(linhas_a_processar, start=1):
    aviso.atualizar(f"Produto {indice}/{len(linhas_a_processar)} — {mlb}", "#c0392b")
    print(f"\n[{indice}/{len(linhas_a_processar)}] {produto} (MLB {mlb}, EAN {ean})")

    marca_limpa = sanitizar_nome(str(marca or "SEM_MARCA"))
    produto_limpo = sanitizar_nome(str(produto))
    pasta_produto = PASTA_RAIZ / NOME_EMPRESA_FIXO / marca_limpa / f"{produto_limpo} # {ean}"

    roteiro = achar_proximo_roteiro(pasta_produto)
    if roteiro is None:
        print(f"  [ERRO] Nenhum Roteiro disponível em {pasta_produto} — pulando.")
        resultados.append((numero_linha, produto, "ERRO_SEM_ROTEIRO"))
        continue

    print(f"  Roteiro escolhido: {roteiro.name}")

    url = montar_url(mlb)
    print(f"  Trocando URL para: {url}")
    janela_ml.set_focus()
    send_keys("^l")
    time.sleep(0.3)
    send_keys(url, with_spaces=True)
    send_keys("{ENTER}")

    # * [EXPLICAÇÃO] → CHECKPOINT 1: tela carregou de verdade.
    # O botão "Subir archivo" tem o MESMO nome em qualquer MLB — checar só
    # "existe" pode pegar a página ANTERIOR ainda de pé, no instante de troca.
    # Corrigido: espera inicial maior + exige 2 checagens seguidas positivas
    # (debounce), reduzindo a chance de confirmar um estado transitório.
    print("  [Checkpoint 1] Esperando a página carregar...")
    time.sleep(1.5)  # dá tempo da página antiga sair de vez antes da 1ª checagem
    carregou = False
    checagens_seguidas = 0
    for _ in range(30):
        time.sleep(0.5)
        existe_agora = janela_ml.child_window(title=NOME_BOTAO_UPLOAD, control_type="Button").exists()
        checagens_seguidas = checagens_seguidas + 1 if existe_agora else 0
        if checagens_seguidas >= 2:
            carregou = True
            break

    if not carregou:
        print(f"  [ERRO] Checkpoint 1 falhou — página não carregou. Pulando.")
        resultados.append((numero_linha, produto, "ERRO_PAGINA_NAO_CARREGOU"))
        continue
    print("  [Checkpoint 1] OK — página carregada.")

    botao_enviar = janela_ml.child_window(title=NOME_BOTAO_ENVIAR, control_type="Button")
    botao_upload = janela_ml.child_window(title=NOME_BOTAO_UPLOAD, control_type="Button")

    janelas_antes = listar_janelas_abertas()
    print("  Clicando no botão de upload...")
    botao_upload.click_input()

    handle_dialogo = None
    for _ in range(25):
        time.sleep(0.2)
        novas = listar_janelas_abertas() - janelas_antes
        if novas:
            handle_dialogo = novas.pop()
            break

    if handle_dialogo is None:
        print("  [ERRO] Janela 'Abrir arquivo' não apareceu — pulando.")
        resultados.append((numero_linha, produto, "ERRO_DIALOGO_NAO_ABRIU"))
        continue

    app_dialogo = Application(backend="uia").connect(handle=handle_dialogo)
    janela_dialogo = app_dialogo.window(handle=handle_dialogo)
    janela_dialogo.child_window(auto_id="1148", control_type="Edit").set_text(str(roteiro))
    janela_dialogo.child_window(auto_id="1", control_type="SplitButton").click_input()

    # * [EXPLICAÇÃO] → CHECKPOINT 2: confirma que o ARQUIVO CERTO foi escolhido —
    # não só que "algum" upload aconteceu. Depois do upload, a área passa a
    # mostrar um botão cujo nome contém o nome do arquivo (ex: "Film camara
    # icon Roteiro 01.mp4"). Se o nome do Roteiro escolhido não aparecer ali,
    # o upload não pegou o arquivo certo (ou não pegou nada).
    print(f"  [Checkpoint 2] Confirmando que '{roteiro.name}' foi escolhido...")
    arquivo_confirmado = False
    for _ in range(20):
        time.sleep(0.5)
        elementos = janela_ml.descendants(control_type="Button")
        if any(roteiro.name in (elem.window_text() or "") for elem in elementos):
            arquivo_confirmado = True
            break

    if not arquivo_confirmado:
        print(f"  [ERRO] Checkpoint 2 falhou — arquivo '{roteiro.name}' não foi confirmado na tela. Pulando.")
        resultados.append((numero_linha, produto, "ERRO_ARQUIVO_NAO_CONFIRMADO"))
        continue
    print("  [Checkpoint 2] OK — arquivo correto confirmado na tela.")

    # * [EXPLICAÇÃO] → CHECKPOINT 3: vídeo processado, pronto pra enviar.
    print("  [Checkpoint 3] Esperando o vídeo processar...")
    habilitou = False
    for _ in range(30):
        time.sleep(0.5)
        if botao_enviar.is_enabled():
            habilitou = True
            break

    if not habilitou:
        print(f"  [ERRO] Checkpoint 3 falhou — vídeo não processou a tempo. Pulando.")
        resultados.append((numero_linha, produto, "ERRO_VIDEO_NAO_PROCESSOU"))
        continue
    print("  [Checkpoint 3] OK — vídeo pronto, botão 'Anunciar' habilitado.")

    # * [EXPLICAÇÃO] → SIMULAÇÃO: não clica em "Anunciar" de verdade ainda.
    print(f"  [SIMULADO] Enviaria o clip agora para {mlb}.")

    novo_nome = roteiro.with_name(f"{PREFIXO_USADO}{roteiro.name}")
    roteiro.rename(novo_nome)
    print(f"  Renomeado para: {novo_nome.name}")

    ws.cell(row=numero_linha, column=9).value = STATUS_SAIDA
    resultados.append((numero_linha, produto, "OK_SIMULADO"))

wb.save(CAMINHO_PLANILHA)
print(f"\nPlanilha salva: {CAMINHO_PLANILHA}")

print("\n=== RESUMO ===")
for numero_linha, produto, situacao in resultados:
    print(f"  Linha {numero_linha} — {produto}: {situacao}")

aviso.atualizar("SIMULAÇÃO CONCLUÍDA", "#2e7d32")
aviso.fechar()
time.sleep(0.3)

root_final = tk.Tk()
root_final.withdraw()
root_final.attributes("-topmost", True)
messagebox.showinfo(
    "Concluído",
    f"Simulação do dia concluída.\n\n{len(resultados)} produto(s) processado(s).\n"
    "Confira o resumo no terminal.",
)
root_final.destroy()

keyboard.unhook_all()
sys.exit(0)