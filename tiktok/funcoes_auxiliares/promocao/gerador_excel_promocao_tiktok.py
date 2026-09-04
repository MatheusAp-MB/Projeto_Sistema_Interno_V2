# tiktok/funcoes_auxiliares/promocao/gerador_excel_promocao_tiktok.py

# Função Objetivo: Gera os 2 arquivos por marca, no formato "FixedPriceWithSKU" do
# TikTok — Preço da oferta direto (já é o nosso "Por", sem cálculo nenhum).

import io
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

# * [EXPLICAÇÃO] → Cabeçalho EXATO do modelo oficial FixedPriceWithSKU do TikTok Shop
#                  (conferido célula a célula, 22/07) — texto multi-linha com as regras
#                  de validação embutidas, idêntico ao arquivo que o TikTok aceita. NUNCA
#                  editar este texto sem conferir de novo contra um modelo oficial atual —
#                  qualquer divergência (mesmo 1 caractere) já causou rejeição de upload.
CABECALHO_PROMOCAO = [
    'Product_id (obrigatório)',
    'SKU_id (obrigatório)',
    'Preço da oferta (obrigatório)\n1. 0 < Preço da oferta ≤ Preço mais baixo nos últimos 30 dias\n2. Preço da oferta < Preço original',
    'Limite total de compra (opcional)\n1. Limite total de compra ≤ Estoque\n2. Em branco refere -se a nenhum limite',
    'Limite de compra do comprador (opcional)\n1. 1 ≤ Limite de compra do comprador ≤ 99\n2. Em branco refere -se a nenhum limite',
]

LABEL_TIPO = {'sem_afiliado': 'Sem Afiliado', 'com_afiliado': 'Com Afiliado'}


# Função Objetivo: Gera o arquivo de subida — só as 5 colunas oficiais, nada além disso.
# Explicação em detalhe: modelo oficial do TikTok Shop (FixedPriceWithSKU) só aceita
# exatamente estas 5 colunas — qualquer coluna extra (mesmo que só informativa) já foi
# confirmado que causa rejeição do upload ("arquivo corrompido"/fora do modelo, 22/07).
# As colunas de conferência (SKU/Produto/Tipo/Estoque) saíram daqui e viraram uma aba
# separada em gerar_excel_detalhes — nunca devem voltar pra este arquivo.
def gerar_excel_promocao(resultados):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    for coluna, texto in enumerate(CABECALHO_PROMOCAO, start=1):
        celula = ws.cell(row=1, column=coluna, value=texto)
        celula.font = Font(name='Calibri', size=10, bold=True)
        celula.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[1].height = 63

    # * [EXPLICAÇÃO] → Mesma validação do modelo oficial: D e E são número inteiro,
    #                  D entre 1 e 99999999, E entre 1 e 99 — ambas permitem célula
    #                  em branco (= "nenhum limite", conforme o próprio cabeçalho diz).
    validacao_d = DataValidation(type='whole', operator='between', formula1='1', formula2='99999999', allow_blank=True)
    validacao_e = DataValidation(type='whole', operator='between', formula1='1', formula2='99', allow_blank=True)
    ws.add_data_validation(validacao_d)
    ws.add_data_validation(validacao_e)
    validacao_d.add('D2:D1048576')
    validacao_e.add('E2:E1048576')

    prontos = [r for r in resultados if r.categoria == 'pronto']
    prontos_ordenados = sorted(prontos, key=lambda r: r.estoque_sistema)

    for indice, r in enumerate(prontos_ordenados, start=2):
        la = r.linha_arquivo
        # * [EXPLICAÇÃO] → Product_id/SKU_id como TEXTO (número_format '@'), igual ao
        #                  modelo oficial — evita notação científica/perda de zero à
        #                  esquerda em IDs longos.
        celula_product = ws.cell(row=indice, column=1, value=str(la.product_id))
        celula_product.number_format = '@'
        celula_sku = ws.cell(row=indice, column=2, value=str(la.sku_id))
        celula_sku.number_format = '@'
        ws.cell(row=indice, column=3, value=float(r.preco_final))
        # Colunas D e E ficam em branco de propósito — "em branco = nenhum limite".

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def gerar_excel_detalhes(resultados):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _adicionar_aba_excecao(wb, 'Preço divergente', [r for r in resultados if r.categoria == 'divergente'])
    _adicionar_aba_excecao(wb, 'Novos (sem Grade)', [r for r in resultados if r.categoria == 'novo'])
    _adicionar_aba_excecao(wb, 'Não encontrados', [r for r in resultados if r.categoria == 'nao_encontrado'])
    _adicionar_aba_excecao(wb, 'Estoque inconsistente', [r for r in resultados if r.categoria == 'estoque_inconsistente'])
    _adicionar_aba_excecao(wb, 'Preço inválido no arquivo', [r for r in resultados if r.categoria == 'preco_invalido'])
    _adicionar_aba_enviados(wb, [r for r in resultados if r.categoria == 'pronto'])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# Função Objetivo: Aba de conferência (SKU/Produto/Tipo/Estoque) dos itens que FORAM
# enviados na promoção — só leitura, nunca sobe na plataforma (fica fora do arquivo
# de subida, que agora é 100% fiel ao modelo oficial, sem colunas extras).
def _adicionar_aba_enviados(wb, resultados):
    ws = wb.create_sheet('Enviados nesta promoção')
    ws.append(['SKU', 'Título', 'Tipo', 'Estoque (sistema)', 'Preço "De" (arquivo)', 'Preço "Por" (enviado)'])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for r in resultados:
        preco_de = r.linha_arquivo.preco_atual if r.linha_arquivo else None
        ws.append([
            r.sku, r.titulo, LABEL_TIPO.get(r.tipo, '—'), r.estoque_sistema,
            float(preco_de) if preco_de is not None else None,
            float(r.preco_final) if r.preco_final is not None else None,
        ])


def _adicionar_aba_excecao(wb, nome_aba, resultados):
    ws = wb.create_sheet(nome_aba)
    ws.append(['SKU', 'Título', 'Tipo', 'Estoque (sistema)', 'Preço plataforma', 'Preço "De" (sistema)'])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for r in resultados:
        preco_plataforma = r.linha_arquivo.preco_atual if r.linha_arquivo else None
        preco_sistema = r.grade.preco_de_exibicao if r.grade else None
        ws.append([
            r.sku, r.titulo, LABEL_TIPO.get(r.tipo, '—'), r.estoque_sistema,
            float(preco_plataforma) if preco_plataforma is not None else None,
            float(preco_sistema) if preco_sistema is not None else None,
        ])