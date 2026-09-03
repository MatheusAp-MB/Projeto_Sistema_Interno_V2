# shopee/funcoes_auxiliares/promocao/gerador_excel_promocao.py

# Função Objetivo: Gera 2 arquivos separados por marca — 1 "seguro pra subir" (só os
# prontos) e 1 "de conferência" (as 5 categorias de exceção juntas).

import io
import openpyxl
from openpyxl.styles import Font


CABECALHO_PROMOCAO = [
    'ID do produto', 'Nome do Produto. (Opcional)', 'Nº de Ref. Parent SKU. (Opcional)',
    'ID de variação', 'Variação de nome. (Opcional)', 'Nº de Ref. SKU. (Opcional)',
    'Preço original (opcional)', 'Preço de desconto', 'Limite de compra (Opcional)',
]


# Função Objetivo: Gera o arquivo PRONTO PRA SUBIR — só a categoria "pronto", 1 aba só.
def gerar_excel_promocao(resultados):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Promoção'
    ws.append(CABECALHO_PROMOCAO)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    prontos = [r for r in resultados if r.categoria == 'pronto']
    # * [EXPLICAÇÃO] → Sem estoque primeiro, com estoque depois — prioridade
    #                  de subida confirmada pelo usuário.
    prontos_ordenados = sorted(prontos, key=lambda r: r.estoque_sistema)

    for r in prontos_ordenados:
        la, g = r.linha_arquivo, r.grade
        # * [EXPLICAÇÃO] → "De": no modo Grade vem do sistema (preco_de_exibicao);
        #                  no modo Arquivo (sem Grade) vem do preço já na plataforma —
        #                  mesma lógica de referência usada no TikTok.
        preco_de = g.preco_de_exibicao if g else la.preco_atual
        ws.append([
            la.id_produto, r.titulo, '', la.id_variacao, '',
            r.sku, float(preco_de), float(r.preco_final), '',
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# Função Objetivo: Gera o arquivo DE CONFERÊNCIA — as 5 categorias de exceção, 1 aba cada.
def gerar_excel_detalhes(resultados):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _adicionar_aba_excecao(wb, 'Preço divergente', [r for r in resultados if r.categoria == 'divergente'])
    _adicionar_aba_excecao(wb, 'Novos (sem Grade)', [r for r in resultados if r.categoria == 'novo'])
    _adicionar_aba_excecao(wb, 'Não encontrados', [r for r in resultados if r.categoria == 'nao_encontrado'])
    _adicionar_aba_excecao(wb, 'Estoque inconsistente', [r for r in resultados if r.categoria == 'estoque_inconsistente'])
    _adicionar_aba_excecao(wb, 'Preço inválido no arquivo', [r for r in resultados if r.categoria == 'preco_invalido'])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _adicionar_aba_excecao(wb, nome_aba, resultados):
    ws = wb.create_sheet(nome_aba)
    ws.append(['SKU', 'Título', 'Estoque (sistema)', 'Preço plataforma', 'Preço "De" (sistema)'])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for r in resultados:
        preco_plataforma = r.linha_arquivo.preco_atual if r.linha_arquivo else None
        preco_sistema = r.grade.preco_de_exibicao if r.grade else None
        ws.append([
            r.sku, r.titulo, r.estoque_sistema,
            float(preco_plataforma) if preco_plataforma is not None else None,
            float(preco_sistema) if preco_sistema is not None else None,
        ])