"""
Etapa 3 -- preenche a Coluna K (CEST) da planilha real da Samvale, cruzando
NCM (Coluna A) e EAN/GTIN (Coluna C) contra o JSON de retorno da API Sysemp
(mesmo arquivo usado na Etapa 1/2 -- o JSON já vem filtrado, sem EAN repetido).

Regra de preenchimento (3 casos, cobrem 100% das linhas):
  1) EAN da planilha existe no JSON E o NCM bate            -> fundo VERDE claro, CEST preenchido
  2) EAN da planilha existe no JSON mas o NCM é diferente   -> fundo AMARELO claro, CEST preenchido
  3) EAN da planilha NÃO existe no JSON                     -> fundo CINZA claro, CEST em branco

Header da planilha real na linha 9, dados a partir da linha 10.

Rodar localmente:
    python preencher_cest_planilha_samvale.py
"""

import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ---------------------------------------------------------------------------
# CONFIGURAÇÃO -- ajuste os 2 caminhos e o nome da aba antes de rodar
# ---------------------------------------------------------------------------

CAMINHO_JSON = Path(
    "integracao_sysemp/retorno_api/dados_impostos_xml_entrada/samvale/"
    "XML_Manifesto_NF_notas_mais_recentes_por_produto.json"
)

CAMINHO_PLANILHA_ALVO = Path("Tabela_A_Preencher.xlsx")  # <-- ajustar
NOME_ABA_ALVO = None  # None = usa a aba ativa; ou coloque o nome exato da aba

LINHA_HEADER = 9
LINHA_INICIO_DADOS = 10
COLUNA_NCM = 1   # A
COLUNA_EAN = 3   # C
COLUNA_CEST = 11  # K

# Chaves reais do JSON (mesmas da Etapa 1/2)
CHAVES_EAN = ["Código Barras"]
CHAVES_NCM = ["NCM Cadastro", "NCM XML"]
CHAVES_CEST = ["CEST"]

_AUSENTE = object()

COR_VERDE = "FFC6EFCE"    # match exato (EAN + NCM batem)
COR_AMARELO = "FFFFEB9C"  # EAN bate, NCM diferente
COR_CINZA = "FFD9D9D9"    # EAN não encontrado no JSON


def normalizar_valor(valor):
    """Mesma normalização da Etapa 1/2: string tem espaços removidos; string
    vazia vira None. Garante que "" e None do JSON sejam tratados igual."""
    if isinstance(valor, str):
        valor = valor.strip()
        return valor or None
    return valor


def normalizar_texto_celula(valor):
    """Normaliza um valor lido de célula do Excel pra comparar com o JSON.
    Se a célula guarda o código como número (Excel às vezes converte código
    de barras/NCM pra número quando não está formatado como texto), remove o
    ".0" residual do float. NÃO faz padding de zero à esquerda -- se um EAN
    perder o zero à esquerda por estar como número na planilha, ele vai cair
    em "não encontrado" e isso aparece no diagnóstico final pra você conferir
    manualmente, em vez do script adivinhar um valor errado."""
    if valor is None:
        return None
    if isinstance(valor, float):
        texto = str(int(valor)) if valor.is_integer() else str(valor)
    else:
        texto = str(valor)
    texto = texto.strip()
    return texto or None


def valor_da_primeira_chave_presente(registro: dict, chaves: list[str]):
    for chave in chaves:
        if chave in registro:
            return chave, registro[chave]
    return None, _AUSENTE


def carregar_registros(caminho: Path) -> list[dict]:
    with caminho.open("r", encoding="utf-8") as arquivo:
        dados = json.load(arquivo)
    if isinstance(dados, list):
        return dados
    if isinstance(dados, dict):
        for chave_candidata in ("registros", "dados", "itens"):
            if isinstance(dados.get(chave_candidata), list):
                return dados[chave_candidata]
        raise ValueError(
            "JSON carregado é um dict sem uma lista de registros reconhecível "
            "('registros'/'dados'/'itens'). Confere a estrutura real do arquivo."
        )
    raise ValueError(f"Formato inesperado no JSON: {type(dados)}")


def construir_indice_por_ean(registros: list[dict]) -> dict:
    """Monta EAN -> (ncm, cest). O JSON já vem filtrado sem EAN repetido
    (confirmado) -- mesmo assim, se algum EAN aparecer 2x com dados
    diferentes, avisa no console em vez de sobrescrever silenciosamente."""
    indice = {}
    conflitos = 0

    for registro in registros:
        _, ean_bruto = valor_da_primeira_chave_presente(registro, CHAVES_EAN)
        _, ncm_bruto = valor_da_primeira_chave_presente(registro, CHAVES_NCM)
        _, cest_bruto = valor_da_primeira_chave_presente(registro, CHAVES_CEST)

        ean = normalizar_valor(None if ean_bruto is _AUSENTE else ean_bruto)
        ncm = normalizar_valor(None if ncm_bruto is _AUSENTE else ncm_bruto)
        cest = normalizar_valor(None if cest_bruto is _AUSENTE else cest_bruto)

        if not ean:
            continue

        if ean in indice and indice[ean] != (ncm, cest):
            conflitos += 1
            continue  # mantém a 1ª ocorrência, só sinaliza

        indice[ean] = (ncm, cest)

    if conflitos:
        print(f"ATENÇÃO: {conflitos} EAN(s) apareceram mais de 1 vez no JSON com NCM/CEST diferentes -- confira.")

    return indice


def preencher_planilha(caminho_planilha: Path, indice_por_ean: dict) -> None:
    workbook = load_workbook(caminho_planilha)
    aba = workbook[NOME_ABA_ALVO] if NOME_ABA_ALVO else workbook.active

    preenchimento_verde = PatternFill("solid", fgColor=COR_VERDE)
    preenchimento_amarelo = PatternFill("solid", fgColor=COR_AMARELO)
    preenchimento_cinza = PatternFill("solid", fgColor=COR_CINZA)

    total_linhas = 0
    total_verde = 0
    total_amarelo = 0
    total_cinza = 0

    for linha in range(LINHA_INICIO_DADOS, aba.max_row + 1):
        celula_ncm = aba.cell(row=linha, column=COLUNA_NCM)
        celula_ean = aba.cell(row=linha, column=COLUNA_EAN)
        celula_cest = aba.cell(row=linha, column=COLUNA_CEST)

        ncm_planilha = normalizar_texto_celula(celula_ncm.value)
        ean_planilha = normalizar_texto_celula(celula_ean.value)

        if ean_planilha is None and ncm_planilha is None:
            continue  # linha em branco no fim da planilha -- ignora, não conta

        total_linhas += 1
        encontrado = indice_por_ean.get(ean_planilha) if ean_planilha else None

        if encontrado is None:
            celula_cest.value = None
            celula_cest.fill = preenchimento_cinza
            total_cinza += 1
            continue

        ncm_json, cest_json = encontrado
        celula_cest.value = cest_json

        if ncm_json == ncm_planilha:
            celula_cest.fill = preenchimento_verde
            total_verde += 1
        else:
            celula_cest.fill = preenchimento_amarelo
            total_amarelo += 1

    workbook.save(caminho_planilha)

    print(f"Linhas processadas (a partir da linha {LINHA_INICIO_DADOS}): {total_linhas}")
    print(f"Verde  (EAN + NCM batem):        {total_verde}")
    print(f"Amarelo (EAN bate, NCM diferente): {total_amarelo}")
    print(f"Cinza  (EAN não encontrado):      {total_cinza}")
    print(f"Planilha atualizada em: {caminho_planilha.resolve()}")


def main() -> None:
    registros = carregar_registros(CAMINHO_JSON)
    indice_por_ean = construir_indice_por_ean(registros)
    print(f"EANs distintos indexados a partir do JSON: {len(indice_por_ean)}")

    preencher_planilha(CAMINHO_PLANILHA_ALVO, indice_por_ean)


if __name__ == "__main__":
    main()