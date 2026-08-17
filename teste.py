# teste.py — script de uso único.
#
# Corrige os cabeçalhos das 2 planilhas de produto ERP da Samvale (SV) pra
# baterem exatamente com o que o importador real do projeto
# (importar_produtos_erp.py) espera — esse código lê cada coluna pelo nome
# EXATO (acento e maiúscula/minúscula importam), sem nenhuma tolerância.
#
# Rode 1 vez, antes do Passo 5 (popular_banco) do passo a passo da SV.
# Depois de confirmar que deu certo, pode apagar este arquivo — ele não faz
# parte do projeto, é só uma ferramenta descartável.
#
# Uso: poetry run python teste.py   (rodar na raiz do repositório)

import shutil
from pathlib import Path

import openpyxl

# --- 1. Ajuste os 2 caminhos abaixo se o nome real do arquivo for diferente
ARQUIVOS = [
    Path('Arquivos usados para Popular Banco/Produtos ERP/Relatorio_Todos_Produtos_Ativos_Tela_Cadastro_Produtos_ERP_SV.xlsx'),
    # Path('Arquivos usados para Popular Banco/Produtos ERP/Relatorio_Todos_Produtos_Inativos_Tela_Cadastro_Produtos_ERP_SV.xlsx'),
]

# --- 2. Mapa de renomeação: nome real na SV -> nome que o código exige -----
RENOMEAR = {
    'Código Auxiliar': 'Codigo Auxiliar',
    'Código de Barras': 'Codigo de Barras',
    'Código Fabricante': 'Codigo do Fabricante',
    'Custo Samvale': 'Custo',
    'altura': 'Altura',
    'largura': 'Largura',
    'comprimento': 'Comprimento',
    'peso_bruto': 'Peso Bruto',
    'inativo': 'Inativo',
    'Produto': 'Detalhes do Produto',  # <- PALPITE — ver aviso no final, confirmar pela amostra
}

# --- 3. Colunas que precisam de conferência visual (ambíguas ou sensíveis) -
COLUNAS_PARA_AMOSTRAR = ['Produto', 'Detalhe do produto', 'inativo']


def fazer_backup(caminho):
    backup = caminho.with_name(caminho.stem + '_BACKUP_ANTES_DO_RENAME' + caminho.suffix)
    if not backup.exists():
        shutil.copy2(caminho, backup)
        print(f'  Backup criado: {backup.name}')
    else:
        print(f'  Backup já existia, não sobrescrito: {backup.name}')


def processar_arquivo(caminho):
    print(f'\n{"=" * 70}')
    print(f'Arquivo: {caminho.name}')
    print(f'{"=" * 70}')

    if not caminho.exists():
        print('  [ERRO] Arquivo não encontrado nesse caminho. Pulando.')
        print(f'         Caminho tentado: {caminho.resolve()}')
        return

    fazer_backup(caminho)

    wb = openpyxl.load_workbook(caminho)
    aba = wb.active

    cabecalho = [celula.value for celula in aba[1]]
    print(f'  Colunas encontradas: {len(cabecalho)}')

    # Mostra amostra real das colunas sensíveis, ANTES de renomear qualquer coisa
    for nome_coluna in COLUNAS_PARA_AMOSTRAR:
        if nome_coluna in cabecalho:
            idx = cabecalho.index(nome_coluna) + 1  # openpyxl é 1-indexado
            amostras = [aba.cell(row=linha, column=idx).value for linha in range(2, 7)]
            print(f'  [AMOSTRA] Coluna "{nome_coluna}" (5 primeiras linhas de dado): {amostras}')

    # Aplica os renames
    renomeados = []
    for coluna_idx, valor in enumerate(cabecalho, start=1):
        if valor in RENOMEAR:
            novo_nome = RENOMEAR[valor]
            aba.cell(row=1, column=coluna_idx).value = novo_nome
            renomeados.append(f'{valor!r} -> {novo_nome!r}')

    if renomeados:
        print('  Renomeado:')
        for linha in renomeados:
            print(f'    - {linha}')
    else:
        print('  Nenhuma coluna do mapa foi encontrada (já renomeado antes?).')

    wb.save(caminho)
    print('  Salvo.')


if __name__ == '__main__':
    for arquivo in ARQUIVOS:
        processar_arquivo(arquivo)

    print(f'\n{"=" * 70}')
    print('CONFIRA ANTES DE RODAR O popular_banco:')
    print('  1) Olhe a [AMOSTRA] da coluna que virou "Detalhes do Produto".')
    print('     Tem nome completo e descritivo do produto (ex: "TENIS X")?')
    print('     Se parecer errado, me avisa ANTES de importar — é só trocar')
    print('     1 linha no RENOMEAR e rodar de novo (o backup já existe).')
    print('  2) Olhe a [AMOSTRA] da coluna "inativo": no arquivo de Ativos')
    print('     deveria ser "F" (ou vazio); no de Inativos, "T". Se a SV usa')
    print('     outro código (ex: "Sim"/"Não"), me avisa que ajusto a regra')
    print('     no importador em vez de mexer no dado.')
    print(f'{"=" * 70}')

