# core/management/commands/popular_banco_suporte/importar_tabela_frete_amazon.py

# Função Objetivo: Importa a tabela de frete da Amazon (DBA + FBA) de uma planilha separada.
# Explicação em detalhe: 2 abas, uma por tipo ("Frete Amazon_DBA"/"Frete Amazon_FBA") — o
# tipo vem do NOME DA ABA, não de uma coluna (mais fácil de editar/conferir na mão). Formato
# LONGO — 1 linha por célula da matriz peso×preço, mais as linhas de faixa baixa
# (peso_min/peso_max vazios = não depende de peso, confirmado na fonte: mesmo valor pra
# qualquer peso abaixo de R$79). Colunas esperadas (linha 1 = cabeçalho):
#   A: Peso mínimo (vazio = faixa de preço baixo, não depende de peso)
#   B: Peso máximo (vazio = sem teto OU não depende de peso, conforme A)
#   C: Preço mínimo
#   D: Preço máximo (vazio = sem teto)
#   E: Valor (R$)

from pathlib import Path
import openpyxl
from precificacao.models import FreteAmazon
from core.funcoes_auxiliares.constantes_performance import BATCH_SIZE_PADRAO
from core.management.commands.popular_banco_suporte.conversor_celula_excel import ConversorCelulaExcel

CAMINHO_TABELA_FRETE_AMAZON = Path('Arquivos_de_Importação/Tabela_Frete_Amazon.xlsx')

ABAS_POR_TIPO = {
    'dba': 'Frete Amazon_DBA',
    'fba': 'Frete Amazon_FBA',
}

COL_PESO_MIN = 0
COL_PESO_MAX = 1
COL_PRECO_MIN = 2
COL_PRECO_MAX = 3
COL_VALOR = 4


class LinhaFreteAmazon:

    def __init__(self, row, tipo, conversor):
        self.row = row
        self.tipo = tipo
        self.conversor = conversor
        self.peso_min = None
        self.peso_max = None
        self.preco_min = None
        self.preco_max = None
        self.valor = None

    def processar(self):
        self.peso_min = self.conversor.para_decimal(self.row[COL_PESO_MIN], casas_decimais=3)
        self.peso_max = self.conversor.para_decimal(self.row[COL_PESO_MAX], casas_decimais=3)
        self.preco_min = self.conversor.para_decimal(self.row[COL_PRECO_MIN], casas_decimais=2)
        self.preco_max = self.conversor.para_decimal(self.row[COL_PRECO_MAX], casas_decimais=2)
        self.valor = self.conversor.para_decimal(self.row[COL_VALOR], casas_decimais=2)
        return self


class ImportadorFreteAmazon:

    def __init__(self, caminho, stdout):
        self.caminho = caminho
        self.stdout = stdout
        self.conversor = ConversorCelulaExcel(origem='openpyxl')

        self.existentes = {}
        self.para_criar = []
        self.para_atualizar = []
        self.erros = []

    def carregar_existentes(self):
        self.existentes = {
            (f.tipo, f.peso_min, f.preco_min): f for f in FreteAmazon.objects.all()
        }

    def processar_aba(self, ws, tipo, indice_erro_base):
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        total = len(rows)

        for indice, row in enumerate(rows, start=1):
            if indice % 25 == 0 or indice == total:
                self.stdout.write(f'    [{tipo.upper()}] ... {indice}/{total} linhas processadas')

            if not any(v is not None for v in row[:5]):
                continue

            try:
                linha = LinhaFreteAmazon(row, tipo, self.conversor).processar()
                if linha.preco_min is None:
                    self.erros.append(f'  [ERRO] {tipo.upper()} linha {indice + 1}: preco_min vazio — corrija a fonte.')
                    continue
                self._registrar_linha(linha)
            except Exception as e:
                self.erros.append(f'  [ERRO] {tipo.upper()} linha {indice + 1}: {e}')

    def _registrar_linha(self, linha):
        chave = (linha.tipo, linha.peso_min, linha.preco_min)
        existente = self.existentes.get(chave)
        if existente and existente.pk:
            existente.peso_max = linha.peso_max
            existente.preco_max = linha.preco_max
            existente.valor = linha.valor
            self.para_atualizar.append(existente)
        else:
            novo = FreteAmazon(
                tipo=linha.tipo, peso_min=linha.peso_min, peso_max=linha.peso_max,
                preco_min=linha.preco_min, preco_max=linha.preco_max, valor=linha.valor,
            )
            self.para_criar.append(novo)
            self.existentes[chave] = novo

    def salvar(self):
        if self.para_criar:
            FreteAmazon.objects.bulk_create(self.para_criar, batch_size=BATCH_SIZE_PADRAO)
        if self.para_atualizar:
            FreteAmazon.objects.bulk_update(
                self.para_atualizar, ['peso_max', 'preco_max', 'valor'], batch_size=BATCH_SIZE_PADRAO,
            )

    def rodar_importacao_completa(self):
        wb = openpyxl.load_workbook(self.caminho, read_only=True, data_only=True)
        self.carregar_existentes()

        for tipo, nome_aba in ABAS_POR_TIPO.items():
            if nome_aba not in wb.sheetnames:
                self.erros.append(f'  [ERRO] Aba "{nome_aba}" não encontrada no arquivo.')
                continue
            self.processar_aba(wb[nome_aba], tipo, indice_erro_base=0)

        self.salvar()
        return self

    def relatorio(self):
        return (
            f'[FRETE AMAZON] Concluído!\n'
            f'    Criados:     {len(self.para_criar)}\n'
            f'    Atualizados: {len(self.para_atualizar)}\n'
            f'    Erros:       {len(self.erros)}'
        )


def importar_tabela_frete_amazon(stdout, style, caminho=CAMINHO_TABELA_FRETE_AMAZON):
    if not caminho.exists():
        stdout.write(f'[FRETE AMAZON] Arquivo {caminho} não encontrado — pulando essa etapa.')
        return

    stdout.write(f'[FRETE AMAZON] Lendo {caminho}...')

    importador = ImportadorFreteAmazon(caminho, stdout)
    importador.rodar_importacao_completa()

    for erro in importador.erros:
        stdout.write(style.ERROR(erro))

    stdout.write(style.SUCCESS(importador.relatorio()))