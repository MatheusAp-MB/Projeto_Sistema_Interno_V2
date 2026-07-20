# core/funcoes_auxiliares/leitor_planilha_robusto.py

# Função Objetivo: Abre uma planilha vinda de fonte externa (upload de usuário,
# arquivo exportado por marketplace) de forma tolerante a XML malformado.
# Explicação em detalhe: arquivos exportados por algumas plataformas (confirmado na
# Shopee) geram .xlsx com XML interno levemente inválido — o openpyxl (rigoroso)
# recusa abrir ("Unable to read workbook... invalid XML"), mas o Excel de verdade
# abre sem problema. python-calamine (motor em Rust, bem mais tolerante) resolve
# isso na prática. Usado SÓ pra ler arquivos de fora — os arquivos que o próprio
# sistema gera sempre abrem limpo com openpyxl, não precisam desse fallback.

import openpyxl


def ler_linhas_planilha_robusta(arquivo_em_memoria, linha_cabecalho, primeira_linha_dado):
    """Devolve (cabecalho, linhas_de_dados) — tenta openpyxl primeiro (mesmo
    padrão já usado no resto do projeto); se falhar por XML inválido, cai pro
    python-calamine, normalizando a saída pro mesmo formato de sempre."""
    try:
        arquivo_em_memoria.seek(0)
        wb = openpyxl.load_workbook(arquivo_em_memoria, data_only=True)
        ws = wb.active
        cabecalho = [c.value for c in next(ws.iter_rows(min_row=linha_cabecalho, max_row=linha_cabecalho))]
        linhas = list(ws.iter_rows(min_row=primeira_linha_dado, values_only=True))
        return cabecalho, linhas
    except Exception:
        from python_calamine import CalamineWorkbook

        arquivo_em_memoria.seek(0)
        wb = CalamineWorkbook.from_filelike(arquivo_em_memoria)
        todas_linhas = wb.get_sheet_by_index(0).to_python()

        cabecalho = todas_linhas[linha_cabecalho - 1] if len(todas_linhas) >= linha_cabecalho else []
        linhas = [tuple(linha) for linha in todas_linhas[primeira_linha_dado - 1:]]
        return cabecalho, linhas