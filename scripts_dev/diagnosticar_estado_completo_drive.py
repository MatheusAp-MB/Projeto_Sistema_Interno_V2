# scripts_dev/diagnosticar_estado_completo_drive.py

# Função Objetivo: Log completo do estado real do Drive, exportado pra
# .xlsx — varre tudo numa passada só, reconstrói a árvore marca→ean→
# Videos→usados (reaproveitando escaneador.py), reconhece cada arquivo
# com o MESMO parser da produção (parser.py, nunca reimplementado à
# parte) e cruza o que o NOME promete com o que o Drive diz de verdade
# (mimeType), critério já registrado no vault ("Badge de Aviso Para
# Arquivos Inconsistentes no Drive"). Só leitura — nunca grava nada no
# Drive nem no banco; o único arquivo escrito é a planilha de saída.

import os
import sys
from datetime import datetime


def _adicionar_raiz_do_projeto_ao_path():
    caminho_atual = os.path.dirname(os.path.abspath(__file__))
    while caminho_atual != os.path.dirname(caminho_atual):
        if os.path.exists(os.path.join(caminho_atual, 'manage.py')):
            sys.path.insert(0, caminho_atual)
            return
        caminho_atual = os.path.dirname(caminho_atual)
    raise RuntimeError('Não foi possível encontrar manage.py subindo a partir deste script.')


_adicionar_raiz_do_projeto_ao_path()

import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'projeto_sistema_interno_mb_sv.settings')
django.setup()

from collections import Counter, defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from rich.console import Console
from rich.panel import Panel

from django.conf import settings
from produtos.models import Produto
from agenda_videos.models import SnapshotArquivosDrive
from agenda_videos.funcoes_auxiliares.drive.cliente import obter_servico_drive
from agenda_videos.funcoes_auxiliares.drive.constantes import MIME_PASTA, NOME_PASTA_VIDEOS
from agenda_videos.funcoes_auxiliares.drive.escaneador import montar_arvore_por_ean
from agenda_videos.funcoes_auxiliares.drive.parser import parsear_arquivos_produto

console = Console()

COR_CABECALHO = '1E3A5F'
COR_OK = 'C6EFCE'
COR_DIVERGENTE = 'FFC7CE'
COR_NAO_RECONHECIDO = 'FFEB9C'


# ============================================================
# Varredura — mesmo padrão paginado de escaneador.py, pedindo também
# 'size' (produção não precisa disso; este diagnóstico precisa pra
# cruzar tipo esperado x conteúdo real).
# ============================================================

def _listar_tudo_paginado_com_tamanho(servico):
    todos_os_itens = []
    page_token = None
    while True:
        resultado = servico.files().list(
            q='trashed = false',
            fields='nextPageToken, files(id, name, mimeType, parents, size)',
            pageSize=1000,
            pageToken=page_token,
        ).execute()
        todos_os_itens.extend(resultado.get('files', []))
        page_token = resultado.get('nextPageToken')
        if not page_token:
            break
    return todos_os_itens


def _formatar_tamanho(tamanho_bruto):
    if not tamanho_bruto:
        return 'N/A'
    tamanho_kb = int(tamanho_bruto) / 1024
    if tamanho_kb > 1024:
        return f'{tamanho_kb / 1024:.1f} MB'
    return f'{tamanho_kb:.1f} KB'


# Função Objetivo: critério de consistência tipo x mimeType real —
# registrado no vault, nunca conferido automaticamente até este script.
def _tipo_consistente_com_mimetype(tipo, mime_type):
    if tipo in ('base', 'completo'):
        return bool(mime_type) and mime_type.startswith('video/')
    if tipo == 'roteiro':
        return not (mime_type and mime_type.startswith('video/'))
    return True


# ============================================================
# Helpers de planilha — únicos, reaproveitados em toda aba, pra nunca
# ficar inconsistente entre uma e outra (mesmo padrão visual de
# mercado_livre/funcoes_auxiliares/exportacao_agenda_videos.py).
# ============================================================

def _estilizar_cabecalho(ws, quantidade_colunas):
    for coluna in range(1, quantidade_colunas + 1):
        celula = ws.cell(row=1, column=coluna)
        celula.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        celula.fill = PatternFill('solid', fgColor=COR_CABECALHO)
        celula.alignment = Alignment(horizontal='center', vertical='center')
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 22


def _ajustar_largura_colunas(ws, larguras):
    for coluna, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(coluna)].width = largura


def _nova_aba(wb, titulo, cabecalho, larguras):
    ws = wb.create_sheet(titulo)
    ws.append(cabecalho)
    _estilizar_cabecalho(ws, len(cabecalho))
    _ajustar_largura_colunas(ws, larguras)
    return ws


# ============================================================
# Varredura de verdade
# ============================================================

servico = obter_servico_drive()
todos_os_itens = _listar_tudo_paginado_com_tamanho(servico)
item_por_id = {item['id']: item for item in todos_os_itens}

raiz_id = settings.GOOGLE_DRIVE_PASTA_RAIZ_MAGAZINE
filhos_da_raiz = [item for item in todos_os_itens if raiz_id in item.get('parents', [])]

arvore_por_ean = montar_arvore_por_ean(todos_os_itens, raiz_id)

filhos_de = defaultdict(list)
for item in todos_os_itens:
    for pai_id in item.get('parents', []):
        filhos_de[pai_id].append(item)

eans_sem_pasta_videos = []
for pasta_marca in filhos_de.get(raiz_id, []):
    if pasta_marca['mimeType'] != MIME_PASTA:
        continue
    for pasta_ean in filhos_de.get(pasta_marca['id'], []):
        if pasta_ean['mimeType'] != MIME_PASTA:
            continue
        tem_pasta_videos = any(
            f['name'].lower() == NOME_PASTA_VIDEOS.lower() and f['mimeType'] == MIME_PASTA
            for f in filhos_de.get(pasta_ean['id'], [])
        )
        if not tem_pasta_videos:
            eans_sem_pasta_videos.append((pasta_marca['name'], pasta_ean['name']))

linhas_arquivo = []
contagem_geral = Counter()

for ean, dados in arvore_por_ean.items():
    marca = dados['marca']
    estrutura_videos = parsear_arquivos_produto(marca, ean, dados['arquivos_videos'])
    estrutura_usados = parsear_arquivos_produto(marca, ean, dados['arquivos_usados'])

    for local, estrutura in (('Videos/', estrutura_videos), ('Videos/usados/', estrutura_usados)):
        for fase_obj in (estrutura.simples, estrutura.video_mensal, estrutura.video_trimestral):
            for ocorrencia in fase_obj.ocorrencias:
                for tipo in ('base', 'roteiro', 'completo'):
                    arquivo = getattr(ocorrencia, tipo)
                    if arquivo is None:
                        continue
                    item_completo = item_por_id.get(arquivo.drive_file_id, {})
                    mime_type = item_completo.get('mimeType')
                    consistente = _tipo_consistente_com_mimetype(tipo, mime_type)
                    veredito = 'OK' if consistente else 'TIPO/CONTEUDO DIVERGENTE'
                    contagem_geral['ok' if consistente else 'tipo_divergente'] += 1
                    linhas_arquivo.append({
                        'marca': marca, 'ean': ean, 'local': local, 'nome': arquivo.nome_arquivo,
                        'fase': fase_obj.fase, 'numero': ocorrencia.numero, 'tipo': tipo,
                        'mime_type': mime_type or '(vazio)', 'tamanho': _formatar_tamanho(item_completo.get('size')),
                        'veredito': veredito,
                    })

        for arquivo in estrutura.arquivos_nao_reconhecidos:
            item_completo = item_por_id.get(arquivo.drive_file_id, {})
            contagem_geral['nao_reconhecidos'] += 1
            linhas_arquivo.append({
                'marca': marca, 'ean': ean, 'local': local, 'nome': arquivo.nome_arquivo,
                'fase': None, 'numero': None, 'tipo': None,
                'mime_type': item_completo.get('mimeType') or '(vazio)',
                'tamanho': _formatar_tamanho(item_completo.get('size')), 'veredito': 'NOME NAO RECONHECIDO',
            })

produtos_nao_encontrados = [ean for ean in arvore_por_ean if not Produto.objects.filter(ean=ean).exists()]

# ============================================================
# Painel-resumo — só no terminal, pra conferência rápida
# ============================================================

console.print(Panel.fit(
    f'[bold]Total de itens no Drive:[/bold] {len(todos_os_itens)}\n'
    f'[bold]EANs reconhecidos pela árvore:[/bold] {len(arvore_por_ean)}\n'
    f'[bold]EANs sem "Videos" encontrada:[/bold] {len(eans_sem_pasta_videos)}\n'
    f'[bold]EANs sem Produto no banco:[/bold] {len(produtos_nao_encontrados)}\n'
    f'[bold]Arquivos analisados:[/bold] {len(linhas_arquivo)}   '
    f'[green]OK: {contagem_geral["ok"]}[/green]   '
    f'[red]Divergente: {contagem_geral["tipo_divergente"]}[/red]   '
    f'[red]Não reconhecido: {contagem_geral["nao_reconhecidos"]}[/red]',
    title='Resumo da Varredura do Drive', border_style='cyan',
))

# ============================================================
# Planilha — 5 abas
# ============================================================

wb = openpyxl.Workbook()
wb.remove(wb.active)

ws_resumo = wb.create_sheet('Resumo')
ws_resumo.append(['Métrica', 'Valor'])
_estilizar_cabecalho(ws_resumo, 2)
_ajustar_largura_colunas(ws_resumo, [40, 20])
for rotulo, valor in [
    ('Total de itens no Drive', len(todos_os_itens)),
    ('EANs reconhecidos pela árvore', len(arvore_por_ean)),
    ('EANs sem pasta "Videos" encontrada', len(eans_sem_pasta_videos)),
    ('EANs sem Produto correspondente no banco', len(produtos_nao_encontrados)),
    ('Arquivos analisados (total)', len(linhas_arquivo)),
    ('Arquivos OK', contagem_geral['ok']),
    ('Arquivos com tipo/conteúdo divergente', contagem_geral['tipo_divergente']),
    ('Arquivos com nome não reconhecido', contagem_geral['nao_reconhecidos']),
    ('Gerado em', datetime.now().strftime('%d/%m/%Y %H:%M:%S')),
]:
    ws_resumo.append([rotulo, valor])

ws_raiz = _nova_aba(wb, 'Filhos da Raiz', ['Tipo', 'Nome', 'ID'], [12, 30, 40])
for item in filhos_da_raiz:
    tipo = 'PASTA' if item['mimeType'] == MIME_PASTA else 'arquivo'
    ws_raiz.append([tipo, item['name'], item['id']])
    if tipo != 'PASTA':
        for celula in ws_raiz[ws_raiz.max_row]:
            celula.fill = PatternFill('solid', fgColor=COR_DIVERGENTE)

ws_estrutura = _nova_aba(wb, 'Problemas de Estrutura', ['Marca', 'Pasta (EAN)'], [22, 22])
for marca, ean in eans_sem_pasta_videos:
    ws_estrutura.append([marca, ean])
    for celula in ws_estrutura[ws_estrutura.max_row]:
        celula.fill = PatternFill('solid', fgColor=COR_NAO_RECONHECIDO)

ws_eans = _nova_aba(
    wb, 'Por EAN',
    ['Marca', 'EAN', 'Produto no Banco', 'Snapshot', 'Videos', 'Videos/usados', 'Qtd Problemas'],
    [18, 18, 18, 18, 10, 14, 14],
)
for ean, dados in sorted(arvore_por_ean.items(), key=lambda kv: (kv[1]['marca'], kv[0])):
    produto = Produto.objects.filter(ean=ean).first()
    snapshot = SnapshotArquivosDrive.objects.filter(produto__ean=ean).first() if produto else None
    problemas_deste_ean = [l for l in linhas_arquivo if l['ean'] == ean and l['veredito'] != 'OK']

    ws_eans.append([
        dados['marca'], ean,
        'Encontrado' if produto else 'NÃO ENCONTRADO',
        snapshot.atualizado_em.strftime('%d/%m/%Y %H:%M') if snapshot else 'sem snapshot',
        len(dados['arquivos_videos']), len(dados['arquivos_usados']), len(problemas_deste_ean),
    ])
    linha_atual = ws_eans.max_row
    if not produto:
        ws_eans.cell(row=linha_atual, column=3).fill = PatternFill('solid', fgColor=COR_DIVERGENTE)
    ws_eans.cell(row=linha_atual, column=7).fill = PatternFill(
        'solid', fgColor=COR_OK if not problemas_deste_ean else COR_DIVERGENTE,
    )

ws_arquivos = _nova_aba(
    wb, 'Detalhe por Arquivo',
    ['Marca', 'EAN', 'Local', 'Arquivo', 'Fase', 'Nº', 'Tipo', 'mimeType Real', 'Tamanho', 'Veredito'],
    [18, 18, 14, 42, 16, 6, 10, 32, 12, 26],
)
cor_por_veredito = {
    'OK': COR_OK, 'TIPO/CONTEUDO DIVERGENTE': COR_DIVERGENTE, 'NOME NAO RECONHECIDO': COR_NAO_RECONHECIDO,
}
for linha in sorted(linhas_arquivo, key=lambda l: (l['marca'], l['ean'], l['local'], l['nome'])):
    ws_arquivos.append([
        linha['marca'], linha['ean'], linha['local'], linha['nome'],
        linha['fase'] or '-', linha['numero'] or '-', linha['tipo'] or '-',
        linha['mime_type'], linha['tamanho'], linha['veredito'],
    ])
    celula_veredito = ws_arquivos.cell(row=ws_arquivos.max_row, column=10)
    celula_veredito.fill = PatternFill('solid', fgColor=cor_por_veredito[linha['veredito']])

caminho_saida = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    f'estado_drive_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
)
wb.save(caminho_saida)
console.print(f'\n[bold green]Planilha salva em:[/bold green] {caminho_saida}')