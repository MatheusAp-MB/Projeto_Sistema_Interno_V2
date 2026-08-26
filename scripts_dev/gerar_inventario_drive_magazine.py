# scripts_dev/gerar_inventario_drive_magazine.py

# Função Objetivo: Gera, numa única execução por empresa (1 única varredura
# do Drive, 1 única consulta ao banco): a planilha detalhada (por arquivo)
# + 3 planilhas de classificação por produto, respondendo 3 perguntas de
# negócio (Plan 1/2/3). Total = 4 planilhas por empresa, salvas dentro de
# "Relatorios DRIVE/<Empresa>/" (pasta criada automaticamente se não
# existir).
#   Plan 1 -> Produto está sendo trabalhado no Drive, com estrutura e
#             nomenclatura corretas.
#   Plan 2 -> Produto está sendo trabalhado no Drive, mas com algum
#             problema de estrutura (pasta vazia/fantasma, nome errado,
#             arquivo fora do lugar, marca inválida, ou EAN duplicado em
#             mais de 1 marca).
#   Plan 3 -> Produto NÃO está sendo trabalhado no Drive — a pasta do EAN
#             nunca foi criada.
# Universo = todo produto ATIVO do banco DA EMPRESA RODADA (não só o que a
# varredura achou) — um produto sem NENHUMA pasta no Drive também aparece,
# no Plan 3. Nada é filtrado da planilha detalhada — as planilhas de
# classificação são uma CLASSIFICAÇÃO em cima do mesmo dado, nunca uma
# exclusão dele. Só leitura — nunca grava nada no Drive nem no banco.
#
# * [EXPLICAÇÃO] → NOVO em 26/08/2026: roda pra Magazine E/OU Samvale (ver
#                  "Como rodar" mais abaixo) — antes era só Magazine. Cada
#                  empresa gera seu próprio conjunto de 4 planilhas, numa
#                  subpasta própria ("Relatorios DRIVE/Magazine/" ou
#                  "Relatorios DRIVE/Samvale/"), pra não misturar nem
#                  sobrescrever os arquivos de uma empresa com os da outra.
# * [EXPLICAÇÃO] → Casamento produto-do-banco x Drive é feito só pelo EAN
#                  (chave única, confirmado pelo usuário) — não importa
#                  embaixo de qual pasta de marca o EAN foi encontrado.
# * [EXPLICAÇÃO] → REVISADO em 26/08/2026 (decisão explícita do usuário,
#                  substitui o critério anterior de 25/08/2026): este
#                  script NÃO valida mais o TIPO/CONTEÚDO do arquivo (se é
#                  vídeo real, roteiro, ou vídeo já usado em "usados/") —
#                  só valida se está SENDO TRABALHADO (existe pelo menos 1
#                  arquivo real, em local válido) e se a NOMENCLATURA desse
#                  arquivo está correta. Um Roteiro sozinho, um vídeo
#                  simples, ou um vídeo dentro de "Videos/usados/" contam
#                  igualmente para Plan 1, desde que o nome siga o padrão
#                  esperado e o arquivo esteja direto em "Videos/" ou em
#                  "Videos/usados/" (nenhuma outra subpasta é válida).
# * [EXPLICAÇÃO] → Pasta do EAN encontrada mas sem NENHUM arquivo real
#                  dentro (seja ela 100% vazia, seja só a pasta "Videos"
#                  criada e vazia) → Plan 2, não Plan 3. Decisão explícita
#                  do usuário (26/08/2026) — isso muda em relação à versão
#                  de 25/08/2026, onde esse caso caía na planilha "Sem
#                  Vídeo" (equivalente ao Plan 3 de hoje). Plan 3 agora é
#                  reservado só para "a pasta do EAN nunca foi criada".
# * [EXPLICAÇÃO] → "Estrutura correta" (Plan 1) também continua exigindo:
#                  nome da pasta de marca válido no banco, e EAN nunca
#                  duplicado em mais de 1 pasta de marca — isso não foi
#                  reconfirmado nesta rodada de decisões (26/08/2026), foi
#                  mantido do critério anterior por não ter sido
#                  contestado. Se não fizer sentido, avisar pra remover.
# * [EXPLICAÇÃO] → Quando mais de 1 problema se aplica ao mesmo produto, o
#                  "Motivo" mostrado no Plan 2 segue esta ordem de
#                  prioridade (só o 1º que bater aparece): pasta vazia >
#                  EAN duplicado em marcas > marca inválida > arquivo fora
#                  do lugar > nome de arquivo inválido.
#
# Como rodar (no ambiente real do projeto, com Drive e banco configurados):
#   python scripts_dev/gerar_inventario_drive_magazine.py            # roda as 2 empresas (Magazine e Samvale)
#   python scripts_dev/gerar_inventario_drive_magazine.py magazine   # roda só a Magazine
#   python scripts_dev/gerar_inventario_drive_magazine.py samvale    # roda só a Samvale

import os
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from typing import Optional


def _adicionar_raiz_do_projeto_ao_path():
    caminho_atual = os.path.dirname(os.path.abspath(__file__))
    while caminho_atual != os.path.dirname(caminho_atual):
        if os.path.exists(os.path.join(caminho_atual, 'manage.py')):
            sys.path.insert(0, caminho_atual)
            return
        caminho_atual = os.path.dirname(caminho_atual)
    raise RuntimeError('Não foi possível encontrar manage.py subindo a partir deste script.')


_adicionar_raiz_do_projeto_ao_path()

import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'projeto_sistema_interno_mb_sv.settings')
django.setup()

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from django.conf import settings
from core.empresa import EMPRESA_MAGAZINE, EMPRESA_SAMVALE, definir_empresa_ativa
from produtos.models import Produto
from agenda_videos.funcoes_auxiliares.drive.cliente import obter_servico_drive
from agenda_videos.funcoes_auxiliares.drive.constantes import MIME_PASTA, NOME_PASTA_USADOS, NOME_PASTA_VIDEOS
from agenda_videos.funcoes_auxiliares.drive.escaneador import _listar_tudo_paginado
from agenda_videos.funcoes_auxiliares.drive.parser import (
    EXTENSOES_VALIDAS_POR_TIPO, PADRAO_NUMERADO, PADRAO_SIMPLES, _extensao_valida, _normalizar_tipo,
)

COR_CABECALHO = '1E3A5F'
COR_AVISO = 'DCE6F1'
COR_VALIDO = 'C6EFCE'
COR_INVALIDO = 'FFC7CE'
COR_NAO_APLICAVEL = 'F2F2F2'

CABECALHO_DETALHE = [
    'Marca', 'EAN', 'Local', 'Nome do Arquivo',
    'Marca Válida', 'EAN Válido', 'Pasta Videos Válida', 'Nome Válido',
]
LARGURAS_DETALHE = [22, 22, 30, 45, 16, 14, 20, 14]
COLUNAS_DE_VALIDACAO_DETALHE = [5, 6, 7, 8]

CABECALHO_CLASSIFICACAO = ['Marca', 'EAN', 'Motivo']
LARGURAS_CLASSIFICACAO = [22, 22, 55]

MARCADOR_PASTA_VAZIA = '(pasta vazia)'
MARCADOR_RAIZ_DO_EAN = '(raiz do EAN)'
MARCADOR_RAIZ_DA_MARCA = '(raiz da marca)'
MARCADOR_RAIZ_DO_DRIVE = '(raiz do Drive)'
MARCADOR_FORA_DE_EAN = '(fora de qualquer EAN)'
MARCADOR_FORA_DE_MARCA = '(fora de qualquer marca)'
MARCADORES_DE_RAIZ = (MARCADOR_RAIZ_DO_EAN, MARCADOR_RAIZ_DA_MARCA, MARCADOR_RAIZ_DO_DRIVE)

MOTIVO_ESTRUTURA_VALIDA = 'Estrutura válida'
MOTIVO_EAN_NAO_ENCONTRADO = 'EAN não encontrado no Drive'
MOTIVO_PASTA_EAN_VAZIA = 'Pasta do EAN existe mas está totalmente vazia (pasta fantasma)'
MOTIVO_SEM_ARQUIVO_REAL = 'Pasta(s) foram criadas, mas nenhum arquivo real dentro'
MOTIVO_EAN_DUPLICADO_EM_MARCAS = 'EAN encontrado em mais de 1 pasta de marca'
MOTIVO_MARCA_INVALIDA = 'Pasta de marca não corresponde a uma marca válida no banco'
MOTIVO_LOCAL_INVALIDO = 'Arquivo fora do lugar esperado (fora de "Videos/" ou "Videos/usados/")'
MOTIVO_NOME_INVALIDO = 'Nome de arquivo não segue o padrão esperado'

# Onde as planilhas de cada empresa são salvas: "Relatorios DRIVE/<nome>/".
PASTA_RELATORIOS_BASE = 'Relatorios DRIVE'
NOME_PASTA_SAIDA_POR_EMPRESA = {
    EMPRESA_MAGAZINE: 'Magazine',
    EMPRESA_SAMVALE: 'Samvale',
}


@dataclass(frozen=True)
class ItemEncontradoNoDrive:
    marca: str
    ean: str
    local: str
    nome: str


# Função Objetivo: Representa o resultado da classificação de 1 produto do
# banco no Plan 1/2/3 (ver explicação no cabeçalho do arquivo).
@dataclass(frozen=True)
class ClassificacaoProduto:
    marca: str
    ean: str
    plano: int
    motivo: str


# ============================================================
# Helpers de planilha.
# ============================================================

def _estilizar_cabecalho(ws: Worksheet, linha: int, quantidade_colunas: int) -> None:
    for coluna in range(1, quantidade_colunas + 1):
        celula = ws.cell(row=linha, column=coluna)
        celula.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        celula.fill = PatternFill('solid', fgColor=COR_CABECALHO)
        celula.alignment = Alignment(horizontal='center', vertical='center')
    ws.freeze_panes = f'A{linha + 1}'
    ws.auto_filter.ref = f'A{linha}:{get_column_letter(quantidade_colunas)}{linha}'
    ws.row_dimensions[linha].height = 22


def _ajustar_largura_colunas(ws: Worksheet, larguras: list) -> None:
    for coluna, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(coluna)].width = largura


def _escrever_aviso(ws: Worksheet, quantidade_colunas: int, texto: str) -> None:
    celula = ws.cell(row=1, column=1, value=texto)
    celula.font = Font(name='Arial', size=9, italic=True, color='31859B')
    celula.fill = PatternFill('solid', fgColor=COR_AVISO)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=quantidade_colunas)


def _sanitizar_nome_de_aba(nome_marca: str) -> str:
    nome_limpo = nome_marca
    for caractere_proibido in [':', '\\', '/', '?', '*', '[', ']']:
        nome_limpo = nome_limpo.replace(caractere_proibido, '-')
    return nome_limpo[:31]


def _cor_para_valor_validacao(valor: str) -> str:
    if valor == 'Sim':
        return COR_VALIDO
    if valor == 'Não':
        return COR_INVALIDO
    return COR_NAO_APLICAVEL


# ============================================================
# Varredura do Drive — sem mudança.
# ============================================================

def _construir_filhos_de(todos_os_itens: list) -> dict:
    filhos_de = defaultdict(list)
    for item in todos_os_itens:
        for pai_id in item.get('parents', []):
            filhos_de[pai_id].append(item)
    return filhos_de


def _varrer_recursivo(filhos_de: dict, pasta_id: str, caminho_relativo: str = '') -> list:
    filhos = filhos_de.get(pasta_id, [])
    if not filhos:
        return [(caminho_relativo, MARCADOR_PASTA_VAZIA)]

    entradas = []
    for filho in sorted(filhos, key=lambda f: f['name'].lower()):
        if filho['mimeType'] == MIME_PASTA:
            entradas.extend(_varrer_recursivo(filhos_de, filho['id'], f"{caminho_relativo}{filho['name']}/"))
        else:
            entradas.append((caminho_relativo, filho['name']))
    return entradas


def _montar_itens_do_drive(todos_os_itens: list, raiz_id: str) -> list:
    filhos_de = _construir_filhos_de(todos_os_itens)
    itens = []

    filhos_da_raiz = filhos_de.get(raiz_id, [])
    for item in filhos_da_raiz:
        if item['mimeType'] != MIME_PASTA:
            itens.append(ItemEncontradoNoDrive(MARCADOR_FORA_DE_MARCA, MARCADOR_FORA_DE_EAN, MARCADOR_RAIZ_DO_DRIVE, item['name']))

    for pasta_marca in filhos_da_raiz:
        if pasta_marca['mimeType'] != MIME_PASTA:
            continue
        marca = pasta_marca['name']
        filhos_da_marca = filhos_de.get(pasta_marca['id'], [])

        for item in filhos_da_marca:
            if item['mimeType'] != MIME_PASTA:
                itens.append(ItemEncontradoNoDrive(marca, MARCADOR_FORA_DE_EAN, MARCADOR_RAIZ_DA_MARCA, item['name']))

        for pasta_ean in filhos_da_marca:
            if pasta_ean['mimeType'] != MIME_PASTA:
                continue
            ean = pasta_ean['name']
            for caminho, nome in _varrer_recursivo(filhos_de, pasta_ean['id']):
                local = caminho if caminho else MARCADOR_RAIZ_DO_EAN
                itens.append(ItemEncontradoNoDrive(marca, ean, local, nome))

    return itens


# ============================================================
# Validações unitárias — reaproveitadas tanto pela planilha detalhada
# quanto pela classificação por produto.
# ============================================================

def _nome_de_arquivo_valido(nome_arquivo: str) -> bool:
    match_simples = PADRAO_SIMPLES.match(nome_arquivo)
    if match_simples:
        tipo, extensao = _normalizar_tipo(match_simples.group(1).lower()), match_simples.group(2)
        return _extensao_valida(tipo, extensao)

    match_numerado = PADRAO_NUMERADO.match(nome_arquivo)
    if match_numerado:
        _, _, tipo, extensao = match_numerado.groups()
        return _extensao_valida(_normalizar_tipo(tipo.lower()), extensao)

    return False


def _validar_nome(nome: str) -> str:
    if nome == MARCADOR_PASTA_VAZIA:
        return '-'
    return 'Sim' if _nome_de_arquivo_valido(nome) else 'Não'


def _marca_valida(marca: str, marcas_validas: set) -> str:
    if marca == MARCADOR_FORA_DE_MARCA:
        return '-'
    return 'Sim' if marca.upper().strip() in marcas_validas else 'Não'


def _ean_valido(ean: str, eans_validos: set) -> str:
    if ean == MARCADOR_FORA_DE_EAN:
        return '-'
    return 'Sim' if ean.strip() in eans_validos else 'Não'


def _primeiro_segmento_do_caminho(local: str) -> Optional[str]:
    if local in MARCADORES_DE_RAIZ:
        return None
    return local.split('/')[0]


def _eh_segmento_pasta_videos(segmento: Optional[str]) -> bool:
    return bool(segmento) and segmento.upper().strip() == NOME_PASTA_VIDEOS.upper().strip()


# Função Objetivo: True quando `local` é "Videos/" (direto) OU
# "Videos/usados/" (arquivado) — os 2 únicos lugares onde um arquivo conta
# como parte válida da estrutura pra Plan 1/2. Qualquer outro local (raiz
# do EAN, raiz da marca, ou qualquer outra subpasta dentro de "Videos"
# diferente de "usados") é "arquivo fora do lugar esperado" (Plan 2).
# REVISADO em 26/08/2026 — antes ("_esta_diretamente_na_pasta_videos")
# "usados/" NÃO contava; agora conta, desde que o nome do arquivo esteja
# correto (decisão do usuário: não validamos mais o TIPO do arquivo).
def _eh_local_valido_para_arquivo(local: str) -> bool:
    segmento = _primeiro_segmento_do_caminho(local)
    if not _eh_segmento_pasta_videos(segmento):
        return False
    return local == f'{segmento}/' or local == f'{segmento}/{NOME_PASTA_USADOS}/'


def _pasta_videos_existe_no_grupo(itens_do_grupo: list) -> bool:
    return any(_eh_segmento_pasta_videos(_primeiro_segmento_do_caminho(item.local)) for item in itens_do_grupo)


def _validar_pasta_videos_por_produto(itens: list) -> dict:
    grupos = defaultdict(list)
    for item in itens:
        grupos[(item.marca, item.ean)].append(item)
    return {par: _pasta_videos_existe_no_grupo(itens_do_grupo) for par, itens_do_grupo in grupos.items()}


# Função Objetivo: True quando a pasta do EAN existe no Drive mas está 100%
# vazia (nem "Videos" foi criada) — usado só pra escolher o texto do
# "Motivo" dentro do Plan 2 (distinguir de "pasta(s) criadas mas sem
# arquivo real", ex: "Videos" criada mas vazia).
def _eh_ean_completamente_vazio(itens_do_produto: list) -> bool:
    return (
        len(itens_do_produto) == 1
        and itens_do_produto[0].local == MARCADOR_RAIZ_DO_EAN
        and itens_do_produto[0].nome == MARCADOR_PASTA_VAZIA
    )


# ============================================================
# Classificação por produto — Plan 1/2/3 (ver explicação no cabeçalho).
# ============================================================

# Função Objetivo: Decide o Plan (1/2/3) de 1 produto do banco, cruzando
# com o que foi achado no Drive pelo mesmo EAN. NÃO valida o TIPO do
# arquivo (vídeo real x roteiro x já usado) — só se está sendo trabalhado
# (tem pelo menos 1 arquivo real, em local válido) e se a nomenclatura
# desse arquivo está correta.
def _classificar_estrutura_produto(marca_db: str, ean_db: str, itens_por_ean: dict, marcas_validas: set) -> ClassificacaoProduto:
    itens_do_produto = itens_por_ean.get(ean_db.strip(), [])

    # Plan 3 — a pasta do EAN nunca foi criada no Drive.
    if not itens_do_produto:
        return ClassificacaoProduto(marca_db, ean_db, 3, MOTIVO_EAN_NAO_ENCONTRADO)

    arquivos_reais = [item for item in itens_do_produto if item.nome != MARCADOR_PASTA_VAZIA]

    # Plan 2 — a(s) pasta(s) do EAN existem, mas nenhum arquivo real dentro.
    if not arquivos_reais:
        if _eh_ean_completamente_vazio(itens_do_produto):
            return ClassificacaoProduto(marca_db, ean_db, 2, MOTIVO_PASTA_EAN_VAZIA)
        return ClassificacaoProduto(marca_db, ean_db, 2, MOTIVO_SEM_ARQUIVO_REAL)

    # A partir daqui, tem pelo menos 1 arquivo real — checa cada tipo de
    # inconsistência, na ordem de prioridade descrita no cabeçalho.
    marcas_encontradas = {item.marca for item in itens_do_produto}
    if len(marcas_encontradas) > 1:
        return ClassificacaoProduto(marca_db, ean_db, 2, MOTIVO_EAN_DUPLICADO_EM_MARCAS)

    if not all(_marca_valida(item.marca, marcas_validas) == 'Sim' for item in itens_do_produto):
        return ClassificacaoProduto(marca_db, ean_db, 2, MOTIVO_MARCA_INVALIDA)

    if any(not _eh_local_valido_para_arquivo(item.local) for item in arquivos_reais):
        return ClassificacaoProduto(marca_db, ean_db, 2, MOTIVO_LOCAL_INVALIDO)

    if any(not _nome_de_arquivo_valido(item.nome) for item in arquivos_reais):
        return ClassificacaoProduto(marca_db, ean_db, 2, MOTIVO_NOME_INVALIDO)

    # Plan 1 — tem arquivo real, todos no lugar certo e com nome válido.
    return ClassificacaoProduto(marca_db, ean_db, 1, MOTIVO_ESTRUTURA_VALIDA)


# ============================================================
# Banco — 1 única consulta.
# ============================================================

# Função Objetivo: Consulta o banco DA EMPRESA PASSADA 1 única vez —
# devolve a lista completa de (marca, ean) ativos + os 2 `set` normalizados
# usados pelas checagens "Marca Válida"/"EAN Válido" da planilha detalhada.
# REVISADO em 26/08/2026 — antes era fixo em EMPRESA_MAGAZINE; agora
# recebe a empresa (EMPRESA_MAGAZINE ou EMPRESA_SAMVALE) como parâmetro,
# pra poder rodar pras 2.
def _buscar_produtos_ativos(empresa: str) -> tuple:
    definir_empresa_ativa(empresa)
    produtos_brutos = Produto.objects.filter(ativo_no_erp=True).values_list('marca', 'ean')
    produtos = [(marca or '(sem marca cadastrada)', ean) for marca, ean in produtos_brutos]

    marcas_validas = {marca.upper().strip() for marca, _ in produtos}
    eans_validos = {ean.strip() for _, ean in produtos}
    return produtos, marcas_validas, eans_validos


# ============================================================
# Planilha detalhada (por arquivo) — sem mudança.
# ============================================================

def _montar_linhas_com_espacamento(
    itens: list, validacao_pasta_videos: dict, marcas_validas: set, eans_validos: set,
    filtrar_por_marca: Optional[str] = None,
) -> list:
    itens_ordenados = sorted(itens, key=lambda i: (i.marca, i.ean, i.local, i.nome))

    linhas = []
    par_anterior = None
    for item in itens_ordenados:
        if filtrar_por_marca and item.marca != filtrar_por_marca:
            continue
        par_atual = (item.marca, item.ean)
        if par_anterior is not None and par_atual != par_anterior:
            linhas.append([])
        par_anterior = par_atual

        pasta_videos_valida = '-' if item.ean == MARCADOR_FORA_DE_EAN else ('Sim' if validacao_pasta_videos.get(par_atual, False) else 'Não')
        linhas.append([
            item.marca, item.ean, item.local, item.nome,
            _marca_valida(item.marca, marcas_validas),
            _ean_valido(item.ean, eans_validos),
            pasta_videos_valida,
            _validar_nome(item.nome),
        ])
    return linhas


def _escrever_aba_detalhe(wb, titulo_aba: str, linhas: list, aviso_texto: str) -> Worksheet:
    ws = wb.create_sheet(_sanitizar_nome_de_aba(titulo_aba))
    _escrever_aviso(ws, len(CABECALHO_DETALHE), aviso_texto)
    ws.append(CABECALHO_DETALHE)
    _estilizar_cabecalho(ws, linha=2, quantidade_colunas=len(CABECALHO_DETALHE))
    _ajustar_largura_colunas(ws, LARGURAS_DETALHE)

    for linha in linhas:
        ws.append(linha)
        if not linha:
            continue
        numero_da_linha = ws.max_row
        for indice_coluna in COLUNAS_DE_VALIDACAO_DETALHE:
            valor = linha[indice_coluna - 1]
            ws.cell(row=numero_da_linha, column=indice_coluna).fill = PatternFill('solid', fgColor=_cor_para_valor_validacao(valor))
    return ws


# ============================================================
# Planilhas por Plan (1/2/3) — 1 único formato pros 3, só muda a cor e o
# conjunto de classificações.
# ============================================================

def _escrever_aba_classificacao(wb, titulo_aba: str, classificacoes: list, cor_da_planilha: str, aviso_texto: str) -> Worksheet:
    ws = wb.create_sheet(_sanitizar_nome_de_aba(titulo_aba))
    _escrever_aviso(ws, len(CABECALHO_CLASSIFICACAO), aviso_texto)
    ws.append(CABECALHO_CLASSIFICACAO)
    _estilizar_cabecalho(ws, linha=2, quantidade_colunas=len(CABECALHO_CLASSIFICACAO))
    _ajustar_largura_colunas(ws, LARGURAS_CLASSIFICACAO)

    for classificacao in sorted(classificacoes, key=lambda c: (c.marca, c.ean)):
        ws.append([classificacao.marca, classificacao.ean, classificacao.motivo])
        ws.cell(row=ws.max_row, column=3).fill = PatternFill('solid', fgColor=cor_da_planilha)
    return ws


def _gerar_planilha_por_plano(classificacoes: list, marcas: list, cor_da_planilha: str, caminho_saida: str, aviso_texto: str) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _escrever_aba_classificacao(wb, 'Geral', classificacoes, cor_da_planilha, aviso_texto)
    for marca in marcas:
        classificacoes_da_marca = [c for c in classificacoes if c.marca == marca]
        if classificacoes_da_marca:
            _escrever_aba_classificacao(wb, marca, classificacoes_da_marca, cor_da_planilha, aviso_texto)
    wb.save(caminho_saida)


# ============================================================
# Orquestração — com feedback de progresso em cada etapa.
# ============================================================

def gerar_todas_as_planilhas(empresa: str) -> None:
    nome_saida = NOME_PASTA_SAIDA_POR_EMPRESA[empresa]
    raiz_id = settings.GOOGLE_DRIVE_PASTA_RAIZ_MAGAZINE if empresa == EMPRESA_MAGAZINE else settings.GOOGLE_DRIVE_PASTA_RAIZ_SAMVALE

    pasta_saida = os.path.join(PASTA_RELATORIOS_BASE, nome_saida)
    os.makedirs(pasta_saida, exist_ok=True)

    print(f'Iniciando {nome_saida.upper()} (Drive completo + banco + 4 planilhas)...')

    print(f'Etapa 1/6 — buscando lista completa de itens do Drive ({nome_saida})... pode levar alguns segundos.')
    servico = obter_servico_drive()
    todos_os_itens = _listar_tudo_paginado(servico)
    print(f'  -> {len(todos_os_itens)} item(ns) bruto(s) recebido(s) do Drive.')

    print('Etapa 2/6 — organizando a árvore (marca -> EAN -> arquivos)...')
    itens = _montar_itens_do_drive(todos_os_itens, raiz_id)
    marcas_encontradas_drive = sorted({item.marca for item in itens if item.marca != MARCADOR_FORA_DE_MARCA})
    print(f'  -> {len(itens)} linha(s) de inventário montadas, {len(marcas_encontradas_drive)} marca(s) encontradas no Drive.')

    print(f'Etapa 3/6 — consultando o banco (produtos ativos da {nome_saida})...')
    produtos_ativos, marcas_validas, eans_validos = _buscar_produtos_ativos(empresa)
    print(f'  -> {len(produtos_ativos)} produto(s) ativo(s) encontrados no banco.')

    print('Etapa 4/6 — calculando validações (Pasta Videos, Nome, Marca, EAN) pra planilha detalhada...')
    validacao_pasta_videos = _validar_pasta_videos_por_produto(itens)

    print('Etapa 5/6 — classificando cada produto do banco (Plan 1/2/3)...')
    itens_por_ean = defaultdict(list)
    for item in itens:
        if item.ean != MARCADOR_FORA_DE_EAN:
            itens_por_ean[item.ean.strip()].append(item)

    classificacoes = [_classificar_estrutura_produto(marca, ean, itens_por_ean, marcas_validas) for marca, ean in produtos_ativos]
    contagem_por_plano = Counter(c.plano for c in classificacoes)
    print(
        f'  -> Plan 1 (estrutura válida): {contagem_por_plano[1]} | '
        f'Plan 2 (com problemas de estrutura): {contagem_por_plano[2]} | '
        f'Plan 3 (fora do Drive): {contagem_por_plano[3]}'
    )

    print(f'Etapa 6/6 — gerando as planilhas em "{pasta_saida}"...')
    marcas_db = sorted({marca for marca, _ in produtos_ativos})

    print('  -> Planilha detalhada (inventario_drive.xlsx)...')
    aviso_detalhado = (
        f'Inventário do Drive ({nome_saida}) + validação completa (Marca, EAN, Pasta Videos, Nome de arquivo). '
        f'Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")}.'
    )
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _escrever_aba_detalhe(wb, 'Geral', _montar_linhas_com_espacamento(itens, validacao_pasta_videos, marcas_validas, eans_validos), aviso_detalhado)
    for marca in marcas_encontradas_drive:
        linhas_da_marca = _montar_linhas_com_espacamento(itens, validacao_pasta_videos, marcas_validas, eans_validos, filtrar_por_marca=marca)
        _escrever_aba_detalhe(wb, marca, linhas_da_marca, aviso_detalhado)
    wb.save(os.path.join(pasta_saida, 'inventario_drive.xlsx'))

    planilhas_por_plano = [
        (1, 'produtos_1_estrutura_valida.xlsx', COR_VALIDO,
         'Plan 1 — Produto está sendo trabalhado no Drive, com estrutura e nomenclatura corretas.'),
        (2, 'produtos_2_problemas_de_estrutura.xlsx', COR_INVALIDO,
         'Plan 2 — Produto está sendo trabalhado no Drive, mas com algum problema de estrutura '
         '(pasta vazia/fantasma, nome errado, arquivo fora do lugar, marca inválida, ou EAN duplicado em mais de 1 marca).'),
        (3, 'produtos_3_fora_do_drive.xlsx', COR_INVALIDO,
         'Plan 3 — Produto NÃO está sendo trabalhado no Drive; a pasta do EAN nunca foi criada.'),
    ]
    for numero_plano, nome_arquivo, cor_da_planilha, aviso_texto in planilhas_por_plano:
        print(f'  -> Planilha do Plan {numero_plano} ({nome_arquivo})...')
        classificacoes_deste_plano = [c for c in classificacoes if c.plano == numero_plano]
        aviso_completo = f'{aviso_texto} Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")}.'
        _gerar_planilha_por_plano(classificacoes_deste_plano, marcas_db, cor_da_planilha, os.path.join(pasta_saida, nome_arquivo), aviso_completo)

    print(f'Concluído {nome_saida.upper()} — 4 planilha(s) geradas em "{pasta_saida}".')


EMPRESAS_EXECUTAVEIS_POR_ARGUMENTO = {
    'magazine': EMPRESA_MAGAZINE,
    'samvale': EMPRESA_SAMVALE,
}

if __name__ == '__main__':
    argumento = sys.argv[1].strip().lower() if len(sys.argv) > 1 else None

    if argumento is None:
        empresas_a_rodar = [EMPRESA_MAGAZINE, EMPRESA_SAMVALE]
    elif argumento in EMPRESAS_EXECUTAVEIS_POR_ARGUMENTO:
        empresas_a_rodar = [EMPRESAS_EXECUTAVEIS_POR_ARGUMENTO[argumento]]
    else:
        raise SystemExit(f'Empresa inválida: "{sys.argv[1]}". Use "magazine", "samvale", ou nenhum argumento pra rodar as 2.')

    for empresa in empresas_a_rodar:
        gerar_todas_as_planilhas(empresa)