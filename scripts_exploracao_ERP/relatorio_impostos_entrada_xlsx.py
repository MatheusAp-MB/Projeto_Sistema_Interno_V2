# scripts_exploracao_ERP/relatorio_impostos_entrada_xlsx.py

# Função Objetivo: Gera 1 planilha Excel, 1 linha por produto, com os
# impostos de entrada (XML/Sysemp) já sincronizados no banco — uso único
# e pontual, pedido pelo usuário pra apresentar ao superior. Estrutura
# (10 grupos, 52 colunas) aprovada previamente via mockup
# (mockups/gerar_mockup_relatorio.py). Cada 1 dos 10 grupos tem 1 cor
# própria, do cabeçalho até a última linha de dado (pedido do usuário,
# 14/08/2026) — facilita achar visualmente onde 1 grupo termina e o
# próximo começa, rolando a planilha.
#
# Fonte dos dados — híbrida, por decisão explícita do usuário (14/08/2026):
# as 44 colunas que já têm coluna própria no banco vêm SEMPRE do banco
# (ImpostosECustosXMLEntradaProduto + as 6 tabelas de imposto). As 8
# colunas que nunca foram persistidas (ID Produto Sysemp, Código Auxiliar,
# CFOP XML/Cadastro, Origem da Mercadoria XML/Cadastro, Natureza da
# Operação Cadastro, TES de Saída Cadastro) vêm do json já salvo em disco
# pela sincronização mais recente (XML_Manifesto_NF_notas_mais_recentes_
# por_produto.json), casado por EAN — nenhuma chamada nova à API. Produto
# do banco sem correspondência nesse json fica com "—" nessas 8 colunas
# (pode ter sido sincronizado numa janela incremental anterior, que não
# está mais no json mais recente).
#
# Regra única de cálculo (confirmada com o usuário): Base de Cálculo e
# Valor de QUALQUER imposto (inclusive % FCP/Valor FCP do ICMS ST) são
# sempre por unidade do produto — nunca o total da nota. Alíquota, Redução
# e % FCP são taxas, nunca divididas.

import os
import sys
from datetime import date
from decimal import Decimal


def _adicionar_raiz_do_projeto_ao_path():
    caminho_atual = os.path.dirname(os.path.abspath(__file__))
    while caminho_atual != os.path.dirname(caminho_atual):
        if os.path.exists(os.path.join(caminho_atual, 'manage.py')):
            sys.path.insert(0, caminho_atual)
            return
        caminho_atual = os.path.dirname(caminho_atual)
    raise RuntimeError('Não foi possível encontrar manage.py subindo a partir deste script.')


_adicionar_raiz_do_projeto_ao_path()

import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'projeto_sistema_interno_mb_sv.settings')
django.setup()

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rich.console import Console

from integracao_sysemp.servicos.arquivos_retorno_api import NOME_ARQUIVO_NOTAS_MAIS_RECENTES, ler_json
from integracao_sysemp.servicos.dados_xml_nf import ClassificacaoFiscalItem, IdentificacaoProduto
from produtos.models import Produto

CAMPO_CODIGO_PRODUTO = 'Código Barras'  # chave de junção com o json (EAN)
CAMINHO_SAIDA = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'saidas', 'Relatorio_Impostos_Entrada.xlsx')

LEGENDA = (
    'Base de Cálculo e Valor de qualquer imposto nesta planilha são sempre por unidade do '
    'produto — nunca o total da nota (1 única lógica, sem mistura). Cada grupo de colunas tem '
    'sua própria cor, do cabeçalho até a última linha de dado. Colunas marcadas "—": produto sem '
    'correspondência no json de apoio (sincronização mais recente).'
)

# * [EXPLICAÇÃO] → Fonte única desta estrutura — ordem/nome de grupos e
#                  colunas aqui é exatamente o que vai pro Excel, igual
#                  aprovado no mockup (mockups/gerar_mockup_relatorio.py).
#                  A 3ª posição de cada tupla é a cor BASE do grupo (hex,
#                  sem "#") — as outras 3 variações usadas na planilha
#                  (subcabeçalho, linha ímpar, linha par) são derivadas
#                  dela por _clarear(), nunca cadastradas à mão, pra não
#                  ter 40 hex soltos pra manter consistentes. Cores
#                  escolhidas pra luminância baixa o bastante pra manter
#                  contraste com texto branco em negrito no cabeçalho.
GRUPOS_DE_COLUNAS = [
    ('Identificação da Nota', ['Nota Fiscal', 'Fornecedor', 'Empresa', 'Data de Emissão', 'Data de Entrada'], '1F4E78'),
    ('Identificação do Produto', [
        'ID Produto (Sysemp)', 'Produto', 'Código de Barras', 'Código Auxiliar',
        'Código do Fabricante', 'Quantidade Recebida na Nota',
    ], '0E6655'),
    ('Custos', ['Custo Unitário', 'Custo Total da Nota'], '1B5E20'),
    ('Classificação Fiscal (XML × Cadastro)', [
        'NCM (XML)', 'NCM (Cadastro)', 'CFOP (XML)', 'CFOP (Cadastro)',
        'Origem da Mercadoria (XML)', 'Origem da Mercadoria (Cadastro)',
        'Natureza da Operação (Cadastro)', 'TES de Saída (Cadastro)',
    ], '5B2C6F'),
    ('ICMS', ['CST (XML)', 'CST (Cadastro)', 'Base de Cálculo', 'Alíquota', 'Redução', 'Valor'], '7D5B0A'),
    ('ICMS ST', ['Base de Cálculo', 'Alíquota', 'Redução', 'Valor', '% FCP', 'Valor FCP'], 'A24E0A'),
    ('ICMS Retido', ['Base de Cálculo', 'Valor'], '8B2E12'),
    ('IPI', ['CST (XML)', 'CST (Cadastro)', 'Base de Cálculo', 'Alíquota', 'Valor'], '205E7A'),
    ('PIS', ['CST (XML)', 'CST (Cadastro)', 'Base de Cálculo', 'Alíquota', 'Redução', 'Valor'], '7B241C'),
    ('COFINS', ['CST (XML)', 'CST (Cadastro)', 'Base de Cálculo', 'Alíquota', 'Redução', 'Valor'], '8E2452'),
]
TOTAL_DE_COLUNAS = sum(len(colunas) for _, colunas, _ in GRUPOS_DE_COLUNAS)


# Função Objetivo: Mistura `cor_hex` com branco — fator 0 devolve a cor original, fator 1
# devolve branco puro. Usado pra derivar as 3 variações mais claras de cada cor base (linha 3,
# linha ímpar de dado, linha par de dado) a partir de 1 único hex por grupo.
def _clarear(cor_hex: str, fator: float) -> str:
    r, g, b = int(cor_hex[0:2], 16), int(cor_hex[2:4], 16), int(cor_hex[4:6], 16)
    r = round(r + (255 - r) * fator)
    g = round(g + (255 - g) * fator)
    b = round(b + (255 - b) * fator)
    return f'{r:02X}{g:02X}{b:02X}'


# Função Objetivo: Devolve as posições (1-indexadas) das colunas cujo nome está em `nomes_alvo`,
# na ordem de GRUPOS_DE_COLUNAS — evita hardcodar número de coluna à mão (e ele desviar se
# GRUPOS_DE_COLUNAS mudar). Só é seguro pra nome de coluna que aparece 1 única vez na estrutura
# inteira (ex: "Data de Emissão"); nomes repetidos entre grupos (ex: "Valor") não devem usar isto.
def _indices_das_colunas(nomes_alvo: set) -> tuple:
    indices = []
    indice_atual = 1
    for _, colunas, _ in GRUPOS_DE_COLUNAS:
        for nome_coluna in colunas:
            if nome_coluna in nomes_alvo:
                indices.append(indice_atual)
            indice_atual += 1
    return tuple(indices)


COLUNAS_DE_DATA = _indices_das_colunas({'Data de Emissão', 'Data de Entrada'})
COLUNAS_DO_JSON_DE_APOIO = _indices_das_colunas({
    'ID Produto (Sysemp)', 'Código Auxiliar', 'CFOP (XML)', 'CFOP (Cadastro)',
    'Origem da Mercadoria (XML)', 'Origem da Mercadoria (Cadastro)',
    'Natureza da Operação (Cadastro)', 'TES de Saída (Cadastro)',
})


# Função Objetivo: Devolve, por coluna (1-indexada), o par (tom_impar, tom_par) de fundo pra
# linha de dado — ambos derivados da cor base do grupo daquela coluna, pra dar 1 zebra sutil
# sem perder a identidade de cor do grupo.
def _construir_cores_de_dado_por_coluna() -> list:
    cores = []
    for _, colunas, cor_base in GRUPOS_DE_COLUNAS:
        tom_impar = _clarear(cor_base, 0.90)
        tom_par = _clarear(cor_base, 0.82)
        cores += [(tom_impar, tom_par)] * len(colunas)
    return cores


CORES_DE_DADO_POR_COLUNA = _construir_cores_de_dado_por_coluna()  # índice 0 = coluna 1


# Função Objetivo: Escreve a linha 1 (legenda, mesclada em toda a largura da planilha).
def _escrever_legenda(planilha):
    celula = planilha.cell(row=1, column=1, value=LEGENDA)
    celula.font = Font(italic=True, size=9, color='555555')
    planilha.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_DE_COLUNAS)
    planilha.row_dimensions[1].height = 18


# Função Objetivo: Escreve as linhas 2 (grupo mesclado, cor base do grupo) e 3 (coluna, tom
# mais claro da mesma cor) a partir de GRUPOS_DE_COLUNAS, e configura largura de coluna +
# congelamento de painel (A4).
def _escrever_cabecalho(planilha):
    coluna_atual = 1
    for titulo_grupo, colunas, cor_base in GRUPOS_DE_COLUNAS:
        inicio = coluna_atual
        fim = coluna_atual + len(colunas) - 1
        cor_subcabecalho = _clarear(cor_base, 0.75)

        celula_grupo = planilha.cell(row=2, column=inicio, value=titulo_grupo)
        celula_grupo.font = Font(bold=True, color='FFFFFF', size=11)
        celula_grupo.alignment = Alignment(horizontal='center', vertical='center')
        if fim > inicio:
            planilha.merge_cells(start_row=2, start_column=inicio, end_row=2, end_column=fim)
        for coluna in range(inicio, fim + 1):
            planilha.cell(row=2, column=coluna).fill = PatternFill('solid', fgColor=cor_base)

        for indice, nome_coluna in enumerate(colunas):
            celula_coluna = planilha.cell(row=3, column=inicio + indice, value=nome_coluna)
            celula_coluna.font = Font(bold=True, size=10)
            celula_coluna.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            celula_coluna.fill = PatternFill('solid', fgColor=cor_subcabecalho)

        coluna_atual = fim + 1

    borda = Border(bottom=Side(style='thin', color='AAAAAA'))
    for coluna in range(1, TOTAL_DE_COLUNAS + 1):
        planilha.column_dimensions[get_column_letter(coluna)].width = 15
        planilha.cell(row=3, column=coluna).border = borda
    planilha.row_dimensions[3].height = 32
    planilha.freeze_panes = 'A4'


# Função Objetivo: Converte valor TOTAL DA NOTA em valor POR UNIDADE — decisão do usuário
# (10/08/2026, estendida 14/08/2026 pro Valor FCP): Base de Cálculo e Valor de QUALQUER
# imposto são sempre por unidade, nunca o total da nota. None (nunca 0) se o valor ainda
# não existir ou a quantidade for None/zero — não estima nem mascara silenciosamente.
def _por_unidade(valor_da_nota, quantidade):
    if valor_da_nota is None or not quantidade:
        return None
    return valor_da_nota / quantidade


# Função Objetivo: Monta os 8 valores que só existem no json de apoio (nunca persistidos no
# banco) — "—" em todos se este produto não tiver correspondência no json mais recente.
def _campos_do_json_de_apoio(registro):
    if registro is None:
        return ['—'] * 8
    identificacao_produto = IdentificacaoProduto.a_partir_do_registro(registro)
    classificacao_fiscal = ClassificacaoFiscalItem.a_partir_do_registro(registro)
    return [
        identificacao_produto.id_produto_sysemp,
        identificacao_produto.codigo_auxiliar,
        classificacao_fiscal.cfop_xml,
        classificacao_fiscal.cfop_cadastro,
        classificacao_fiscal.origem_mercadoria_xml,
        classificacao_fiscal.origem_mercadoria_cadastro,
        classificacao_fiscal.natureza_operacao_cadastro,
        classificacao_fiscal.tes_saida_cadastro,
    ]


# Função Objetivo: Monta a linha completa (52 valores, na ordem de GRUPOS_DE_COLUNAS) de 1 produto.
def _montar_linha(produto, registro):
    impostos = produto.impostos_entrada
    quantidade = impostos.quantidade_nota

    valores = [
        impostos.nr_nf, impostos.fornecedor, impostos.empresa_fantasia,
        impostos.emissao, impostos.data_entrada_nota,
    ]

    (
        id_produto_sysemp, codigo_auxiliar, cfop_xml, cfop_cadastro, origem_xml, origem_cadastro,
        natureza_operacao_cadastro, tes_saida_cadastro,
    ) = _campos_do_json_de_apoio(registro)
    valores += [
        id_produto_sysemp, produto.titulo, produto.ean, codigo_auxiliar,
        produto.cod_fabricante, quantidade,
    ]

    valores += [impostos.custo_unitario, impostos.custo_total]

    valores += [
        impostos.ncm_xml, impostos.ncm_cadastro, cfop_xml, cfop_cadastro,
        origem_xml, origem_cadastro, natureza_operacao_cadastro, tes_saida_cadastro,
    ]

    icms = impostos.icms
    valores += [
        icms.cst_xml, icms.cst_cadastro, _por_unidade(icms.base_calculo, quantidade),
        icms.aliquota, icms.reducao, _por_unidade(icms.valor, quantidade),
    ]

    icms_st = impostos.icms_st
    valores += [
        _por_unidade(icms_st.base_calculo, quantidade), icms_st.aliquota, icms_st.reducao,
        _por_unidade(icms_st.valor, quantidade), icms_st.aliquota_fcp,
        _por_unidade(icms_st.valor_fcp, quantidade),
    ]

    icms_ret = impostos.icms_ret
    valores += [_por_unidade(icms_ret.base_calculo, quantidade), _por_unidade(icms_ret.valor, quantidade)]

    ipi = impostos.ipi
    valores += [
        ipi.cst_xml, ipi.cst_cadastro, _por_unidade(ipi.base_calculo, quantidade),
        ipi.aliquota, _por_unidade(ipi.valor, quantidade),
    ]

    pis = impostos.pis
    valores += [
        pis.cst_xml, pis.cst_cadastro, _por_unidade(pis.base_calculo, quantidade),
        pis.aliquota, pis.reducao, _por_unidade(pis.valor, quantidade),
    ]

    cofins = impostos.cofins
    valores += [
        cofins.cst_xml, cofins.cst_cadastro, _por_unidade(cofins.base_calculo, quantidade),
        cofins.aliquota, cofins.reducao, _por_unidade(cofins.valor, quantidade),
    ]

    return valores


def main():
    console = Console()

    console.print('Lendo json de apoio já salvo em disco (sincronização mais recente)...')
    selecionados = ler_json(NOME_ARQUIVO_NOTAS_MAIS_RECENTES, padrao=[])
    registros_por_ean = {registro[CAMPO_CODIGO_PRODUTO]: registro for registro in selecionados}

    console.print('Lendo produtos do banco...')
    produtos = (
        Produto.objects
        .filter(impostos_entrada__isnull=False)
        .select_related(
            'impostos_entrada', 'impostos_entrada__icms', 'impostos_entrada__icms_st',
            'impostos_entrada__icms_ret', 'impostos_entrada__ipi', 'impostos_entrada__pis',
            'impostos_entrada__cofins',
        )
        .order_by('titulo')
    )

    livro = openpyxl.Workbook()
    planilha = livro.active
    planilha.title = 'Impostos de Entrada'
    _escrever_legenda(planilha)
    _escrever_cabecalho(planilha)

    linha_atual = 4
    produtos_sem_json = 0

    for produto in produtos:
        registro = registros_por_ean.get(produto.ean)
        if registro is None:
            produtos_sem_json += 1

        for indice, valor in enumerate(_montar_linha(produto, registro), start=1):
            celula = planilha.cell(row=linha_atual, column=indice, value=valor)

            tom_impar, tom_par = CORES_DE_DADO_POR_COLUNA[indice - 1]
            cor_de_fundo = tom_par if linha_atual % 2 == 0 else tom_impar
            celula.fill = PatternFill('solid', fgColor=cor_de_fundo)

            if indice in COLUNAS_DE_DATA and isinstance(valor, date):
                celula.number_format = 'DD/MM/YYYY'
            if indice in COLUNAS_DO_JSON_DE_APOIO and valor == '—':
                celula.font = Font(italic=True, color='999999')

        linha_atual += 1

    os.makedirs(os.path.dirname(CAMINHO_SAIDA), exist_ok=True)
    livro.save(CAMINHO_SAIDA)

    total_exportado = linha_atual - 4
    console.print(f'[bold green]Planilha gerada:[/bold green] {CAMINHO_SAIDA}')
    console.print(f'Produtos exportados: {total_exportado}')
    if produtos_sem_json:
        console.print(
            f'[yellow]{produtos_sem_json} produto(s) sem correspondência no json de apoio — '
            f'ID Produto (Sysemp)/Código Auxiliar/CFOP/Origem/Natureza da Operação/TES de Saída '
            f'ficaram "—" pra eles (sincronizados numa janela incremental anterior, fora do json '
            f'mais recente).[/yellow]'
        )


if __name__ == '__main__':
    main()