# impostos/funcoes_auxiliares/exportacao_resumo_entrada.py

# Função Objetivo: Gera o .xlsx de Impostos de Entrada (10 grupos, 52
# colunas, 1 cor por grupo do cabeçalho até a última linha de dado) — mesma
# estrutura de scripts_exploracao_ERP/relatorio_impostos_entrada_xlsx.py,
# só que lendo tudo direto do banco (16/08/2026): os 8 campos que antes só
# existiam no json de apoio (ID Produto Sysemp, Código Auxiliar, CFOP,
# Origem, Natureza da Operação, TES de Saída) já são colunas persistidas
# desde a Rodada 2 (15/08/2026) — sem merge de arquivo, sem chamada à API.
# Todo cálculo "por unidade" vem de obter_detalhes_para_exibicao() (mesma
# função que o modal do produto usa) — nunca recalculado aqui de novo.
#
# O script antigo continua existindo, intocado, por decisão do usuário
# (16/08/2026) — mantido só pra rodar local sem precisar do servidor. Isso
# significa que a lista/cor de grupos abaixo existe hoje em 2 lugares; se
# o script antigo for aposentado no futuro, dá pra apontar ele pra cá.
#
# Largura/altura/formato numérico (16/08/2026): nenhum dos 2 geradores
# (este e o script antigo) jamais tratou isso — largura sempre foi
# uniforme (15) e só data tinha number_format. Corrigido aqui: largura por
# NOME de coluna (LARGURA_POR_COLUNA), altura maior nas linhas de dado, e
# number_format por categoria (moeda/percentual/quantidade/texto forçado).
# Os valores em si (Alíquota/Redução) já vêm em UNIDADE DE PORCENTAGEM do
# banco (18.0000 = 18%, não 0.18) — por isso o formato percentual usa
# 0.00"%" (só exibe o símbolo, sem multiplicar por 100), nunca o formato
# nativo do Excel 0.00% (que multiplicaria de novo).

import io
from datetime import date

import openpyxl

from impostos.funcoes_auxiliares.exibicao_impostos_entrada import montar_detalhes_para_exibicao
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

LEGENDA = (
    'Base de Cálculo e Valor de qualquer imposto nesta planilha são sempre por unidade do '
    'produto — nunca o total da nota (1 única lógica, sem mistura). Cada grupo de colunas tem '
    'sua própria cor, do cabeçalho até a última linha de dado. Colunas marcadas "—": produto '
    'ainda não reprocessado desde que esses campos passaram a ser persistidos (15/08/2026).'
)

GRUPOS_DE_COLUNAS = [
    ('Identificação da Nota', ['Nota Fiscal', 'Fornecedor', 'Empresa', 'Data de Emissão', 'Data de Entrada'], '1F4E78'),
    ('Identificação do Produto', [
        'ID Produto (Sysemp)', 'Produto', 'Código de Barras', 'Código Auxiliar',
        'Código do Fabricante', 'Quantidade Recebida na Nota',
    ], '0E6655'),
    ('Custos', ['Custo Unitário', 'Custo Total da Nota'], '1B5E20'),
    ('Classificação Fiscal (XML × Cadastro)', [
        'NCM (XML)', 'NCM (Cadastro)', 'CFOP (XML)', 'CFOP (Cadastro)',
        'Origem da Mercadoria (XML)', 'Origem da Mercadoria (Cadastro)',
        'Natureza da Operação (Cadastro)', 'TES de Saída (Cadastro)',
    ], '5B2C6F'),
    ('ICMS', ['CST (XML)', 'CST (Cadastro)', 'Base de Cálculo', 'Alíquota', 'Redução', 'Valor'], '7D5B0A'),
    ('ICMS ST', ['Base de Cálculo', 'Alíquota', 'Redução', 'Valor', '% FCP', 'Valor FCP'], 'A24E0A'),
    ('ICMS Retido', ['Base de Cálculo', 'Valor'], '8B2E12'),
    ('IPI', ['CST (XML)', 'CST (Cadastro)', 'Base de Cálculo', 'Alíquota', 'Valor'], '205E7A'),
    ('PIS', ['CST (XML)', 'CST (Cadastro)', 'Base de Cálculo', 'Alíquota', 'Redução', 'Valor'], '7B241C'),
    ('COFINS', ['CST (XML)', 'CST (Cadastro)', 'Base de Cálculo', 'Alíquota', 'Redução', 'Valor'], '8E2452'),
]
TOTAL_DE_COLUNAS = sum(len(colunas) for _, colunas, _ in GRUPOS_DE_COLUNAS)

# Largura por NOME de coluna — mesmo nome em grupos diferentes (ex:
# "Valor", "Alíquota") compartilha a mesma largura de propósito, já que é
# o mesmo tipo de dado em qualquer grupo. Nome fora daqui usa LARGURA_PADRAO.
LARGURA_PADRAO = 15
LARGURA_POR_COLUNA = {
    'Nota Fiscal': 12, 'Fornecedor': 42, 'Empresa': 20,
    'Data de Emissão': 13, 'Data de Entrada': 13,
    'ID Produto (Sysemp)': 14, 'Produto': 54, 'Código de Barras': 16,
    'Código Auxiliar': 20, 'Código do Fabricante': 18, 'Quantidade Recebida na Nota': 18,
    'Custo Unitário': 13, 'Custo Total da Nota': 15,
    'NCM (XML)': 12, 'NCM (Cadastro)': 12, 'CFOP (XML)': 10, 'CFOP (Cadastro)': 10,
    'Origem da Mercadoria (XML)': 28, 'Origem da Mercadoria (Cadastro)': 28,
    'Natureza da Operação (Cadastro)': 36, 'TES de Saída (Cadastro)': 14,
    'CST (XML)': 10, 'CST (Cadastro)': 10, 'Base de Cálculo': 13,
    'Alíquota': 10, 'Redução': 10, 'Valor': 12, '% FCP': 9, 'Valor FCP': 12,
}

ALTURA_LINHA_DADO = 36  # espaço suficiente pra Produto/Fornecedor quebrarem em 2 linhas


def _clarear(cor_hex: str, fator: float) -> str:
    r, g, b = int(cor_hex[0:2], 16), int(cor_hex[2:4], 16), int(cor_hex[4:6], 16)
    r = round(r + (255 - r) * fator)
    g = round(g + (255 - g) * fator)
    b = round(b + (255 - b) * fator)
    return f'{r:02X}{g:02X}{b:02X}'


def _indices_das_colunas(nomes_alvo: set) -> tuple:
    indices = []
    indice_atual = 1
    for _, colunas, _ in GRUPOS_DE_COLUNAS:
        for nome_coluna in colunas:
            if nome_coluna in nomes_alvo:
                indices.append(indice_atual)
            indice_atual += 1
    return tuple(indices)


COLUNAS_DE_DATA = _indices_das_colunas({'Data de Emissão', 'Data de Entrada'})
COLUNAS_QUE_PODEM_FALTAR = _indices_das_colunas({
    'ID Produto (Sysemp)', 'Código Auxiliar', 'CFOP (XML)', 'CFOP (Cadastro)',
    'Origem da Mercadoria (XML)', 'Origem da Mercadoria (Cadastro)',
    'Natureza da Operação (Cadastro)', 'TES de Saída (Cadastro)',
})

# Categorias de formato numérico — cada uma vira number_format aplicado
# na hora de escrever a célula de dado (ver gerar_excel_resumo_impostos_entrada).
COLUNAS_MONETARIAS = _indices_das_colunas({
    'Custo Unitário', 'Custo Total da Nota', 'Base de Cálculo', 'Valor', 'Valor FCP',
})
COLUNAS_PERCENTUAL = _indices_das_colunas({'Alíquota', 'Redução', '% FCP'})
COLUNAS_QUANTIDADE = _indices_das_colunas({'Quantidade Recebida na Nota'})
COLUNAS_TEXTO_FORCADO = _indices_das_colunas({
    'Nota Fiscal', 'ID Produto (Sysemp)', 'Código de Barras', 'Código Auxiliar',
    'NCM (XML)', 'NCM (Cadastro)', 'CFOP (XML)', 'CFOP (Cadastro)',
    'TES de Saída (Cadastro)', 'CST (XML)', 'CST (Cadastro)',
})
COLUNAS_COM_QUEBRA_DE_LINHA = _indices_das_colunas({
    'Produto', 'Fornecedor', 'Natureza da Operação (Cadastro)',
})

FORMATO_MOEDA = '"R$" #,##0.00'
FORMATO_PERCENTUAL = '0.00"%"'
FORMATO_QUANTIDADE = '#,##0.000'
FORMATO_TEXTO = '@'

# Borda fina cinza-clara em toda célula (cabeçalho e dado) — acabamento
# de grade, pedido na revisão geral (16/08/2026).
_LADO_BORDA = Side(style='thin', color='D9D9D9')
BORDA_CELULA = Border(left=_LADO_BORDA, right=_LADO_BORDA, top=_LADO_BORDA, bottom=_LADO_BORDA)


def _construir_cores_de_dado_por_coluna() -> list:
    cores = []
    for _, colunas, cor_base in GRUPOS_DE_COLUNAS:
        tom_impar = _clarear(cor_base, 0.90)
        tom_par = _clarear(cor_base, 0.82)
        cores += [(tom_impar, tom_par)] * len(colunas)
    return cores


CORES_DE_DADO_POR_COLUNA = _construir_cores_de_dado_por_coluna()


def _escrever_legenda(planilha):
    celula = planilha.cell(row=1, column=1, value=LEGENDA)
    celula.font = Font(italic=True, size=9, color='555555')
    planilha.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_DE_COLUNAS)
    planilha.row_dimensions[1].height = 18


def _escrever_cabecalho(planilha):
    coluna_atual = 1
    for titulo_grupo, colunas, cor_base in GRUPOS_DE_COLUNAS:
        inicio = coluna_atual
        fim = coluna_atual + len(colunas) - 1
        cor_subcabecalho = _clarear(cor_base, 0.75)

        celula_grupo = planilha.cell(row=2, column=inicio, value=titulo_grupo)
        celula_grupo.font = Font(bold=True, color='FFFFFF', size=11)
        celula_grupo.alignment = Alignment(horizontal='center', vertical='center')
        celula_grupo.border = BORDA_CELULA
        if fim > inicio:
            planilha.merge_cells(start_row=2, start_column=inicio, end_row=2, end_column=fim)
        for coluna in range(inicio, fim + 1):
            planilha.cell(row=2, column=coluna).fill = PatternFill('solid', fgColor=cor_base)
            planilha.cell(row=2, column=coluna).border = BORDA_CELULA

        for indice, nome_coluna in enumerate(colunas):
            celula_coluna = planilha.cell(row=3, column=inicio + indice, value=nome_coluna)
            celula_coluna.font = Font(bold=True, size=10)
            celula_coluna.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            celula_coluna.fill = PatternFill('solid', fgColor=cor_subcabecalho)
            celula_coluna.border = BORDA_CELULA

            largura = LARGURA_POR_COLUNA.get(nome_coluna, LARGURA_PADRAO)
            planilha.column_dimensions[get_column_letter(inicio + indice)].width = largura

        coluna_atual = fim + 1

    planilha.row_dimensions[3].height = 32
    planilha.freeze_panes = 'A4'


# Função Objetivo: Monta a linha completa (52 valores, na ordem de
# GRUPOS_DE_COLUNAS) de 1 produto — tudo vindo de obter_detalhes_para_exibicao()
# (mesma conversão "por unidade" que o modal usa) mais os 3 campos que só
# existem no Produto (titulo, ean, cod_fabricante).
def _montar_linha(produto):
    detalhes = montar_detalhes_para_exibicao(produto.impostos_entrada)
    por_imposto = {linha.nome: linha for linha in detalhes.linhas}

    valores = [
        detalhes.nr_nf, detalhes.fornecedor, detalhes.empresa_fantasia,
        detalhes.emissao, detalhes.data_entrada_nota,
    ]
    valores += [
        detalhes.id_produto_sysemp, produto.titulo, produto.ean, detalhes.codigo_auxiliar,
        produto.cod_fabricante, detalhes.quantidade_nota,
    ]
    valores += [detalhes.custo_unitario, detalhes.custo_total]
    valores += [
        detalhes.ncm_xml, detalhes.ncm_cadastro, detalhes.cfop_xml, detalhes.cfop_cadastro,
        detalhes.origem_mercadoria_xml, detalhes.origem_mercadoria_cadastro,
        detalhes.natureza_operacao_cadastro, detalhes.tes_saida_cadastro,
    ]

    icms = por_imposto['ICMS']
    valores += [icms.cst_xml, icms.cst_cadastro, icms.base_calculo, icms.aliquota, icms.reducao, icms.valor]

    icms_st = por_imposto['ICMS ST']
    valores += [
        icms_st.base_calculo, icms_st.aliquota, icms_st.reducao,
        icms_st.valor, icms_st.aliquota_fcp, icms_st.valor_fcp,
    ]

    icms_ret = por_imposto['ICMS Retido']
    valores += [icms_ret.base_calculo, icms_ret.valor]

    ipi = por_imposto['IPI']
    valores += [ipi.cst_xml, ipi.cst_cadastro, ipi.base_calculo, ipi.aliquota, ipi.valor]

    pis = por_imposto['PIS']
    valores += [pis.cst_xml, pis.cst_cadastro, pis.base_calculo, pis.aliquota, pis.reducao, pis.valor]

    cofins = por_imposto['COFINS']
    valores += [cofins.cst_xml, cofins.cst_cadastro, cofins.base_calculo, cofins.aliquota, cofins.reducao, cofins.valor]

    return valores


# Função Objetivo: Gera o .xlsx completo em memória (bytes), pronto pra
# devolver numa HttpResponse de download — sem escrever em disco.
def gerar_excel_resumo_impostos_entrada(produtos) -> bytes:
    livro = openpyxl.Workbook()
    planilha = livro.active
    planilha.title = 'Impostos de Entrada'
    _escrever_legenda(planilha)
    _escrever_cabecalho(planilha)

    linha_atual = 4
    for produto in produtos:
        planilha.row_dimensions[linha_atual].height = ALTURA_LINHA_DADO

        for indice, valor in enumerate(_montar_linha(produto), start=1):
            eh_campo_ausente = indice in COLUNAS_QUE_PODEM_FALTAR and valor is None
            valor_para_exibir = '—' if eh_campo_ausente else valor
            celula = planilha.cell(row=linha_atual, column=indice, value=valor_para_exibir)

            tom_impar, tom_par = CORES_DE_DADO_POR_COLUNA[indice - 1]
            celula.fill = PatternFill('solid', fgColor=(tom_par if linha_atual % 2 == 0 else tom_impar))
            celula.border = BORDA_CELULA

            if indice in COLUNAS_DE_DATA and isinstance(valor, date):
                celula.number_format = 'DD/MM/YYYY'
            elif indice in COLUNAS_MONETARIAS and valor is not None:
                celula.number_format = FORMATO_MOEDA
            elif indice in COLUNAS_PERCENTUAL and valor is not None:
                celula.number_format = FORMATO_PERCENTUAL
            elif indice in COLUNAS_QUANTIDADE and valor is not None:
                celula.number_format = FORMATO_QUANTIDADE
            elif indice in COLUNAS_TEXTO_FORCADO and not eh_campo_ausente:
                celula.number_format = FORMATO_TEXTO

            # Todo texto centralizado (horizontal e vertical) — pedido na
            # revisão geral (16/08/2026). Colunas de texto longo continuam
            # quebrando linha, só que também centralizadas agora.
            celula.alignment = Alignment(
                horizontal='center', vertical='center',
                wrap_text=indice in COLUNAS_COM_QUEBRA_DE_LINHA,
            )

            if eh_campo_ausente:
                celula.font = Font(italic=True, color='999999')

        linha_atual += 1

    buffer = io.BytesIO()
    livro.save(buffer)
    return buffer.getvalue()